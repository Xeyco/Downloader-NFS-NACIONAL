"""
NFSe Downloader - Versão 16 (Melhorada)
Sistema automatizado de download de Notas Fiscais de Serviço
com melhorias em segurança, performance e tratamento de erros.

Autor: Melhorado por Claude AI
Data: 2026-02-10
"""

import os
import json
import threading
import time
import calendar
import logging
import hashlib
import traceback
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from pathlib import Path
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout
from cryptography.fernet import Fernet
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURAÇÕES GLOBAIS
# ============================================================================

ARQUIVO_CONFIG = 'empresas.json'
ARQUIVO_CHAVE = '.key'
LOG_FILE = 'nfse_downloader.log'
CACHE_FILE = 'downloads_cache.json'

# Configuração de logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


# ============================================================================
# CLASSE DE CRIPTOGRAFIA
# ============================================================================

class CryptoManager:
    """Gerenciador de criptografia para senhas e dados sensíveis"""
    
    def __init__(self):
        self.cipher = None
        self._inicializar_chave()
    
    def _inicializar_chave(self):
        """Inicializa ou carrega a chave de criptografia"""
        if os.path.exists(ARQUIVO_CHAVE):
            with open(ARQUIVO_CHAVE, 'rb') as f:
                chave = f.read()
        else:
            chave = Fernet.generate_key()
            with open(ARQUIVO_CHAVE, 'wb') as f:
                f.write(chave)
            # Tornar arquivo oculto no Windows
            try:
                import ctypes
                ctypes.windll.kernel32.SetFileAttributesW(ARQUIVO_CHAVE, 2)
            except:
                pass
        
        self.cipher = Fernet(chave)
    
    def criptografar(self, texto: str) -> str:
        """Criptografa um texto"""
        if not texto:
            return ""
        return self.cipher.encrypt(texto.encode()).decode()
    
    def descriptografar(self, texto_criptografado: str) -> str:
        """Descriptografa um texto"""
        if not texto_criptografado:
            return ""
        try:
            return self.cipher.decrypt(texto_criptografado.encode()).decode()
        except:
            logger.warning("Erro ao descriptografar. Retornando vazio.")
            return ""


# ============================================================================
# CLASSE DE CACHE
# ============================================================================

class DownloadCache:
    """Gerenciador de cache para evitar downloads duplicados"""
    
    def __init__(self):
        self.cache = self._carregar_cache()
    
    def _carregar_cache(self):
        """Carrega cache de downloads do arquivo"""
        if os.path.exists(CACHE_FILE):
            try:
                with open(CACHE_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                return {}
        return {}
    
    def _salvar_cache(self):
        """Salva cache no arquivo"""
        try:
            with open(CACHE_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.cache, f, indent=2)
        except Exception as e:
            logger.error(f"Erro ao salvar cache: {e}")
    
    def gerar_hash(self, empresa: str, competencia: str, tomador: str) -> str:
        """Gera hash único para identificar uma nota"""
        dados = f"{empresa}_{competencia}_{tomador}"
        return hashlib.md5(dados.encode()).hexdigest()
    
    def ja_baixado(self, hash_nota: str) -> bool:
        """Verifica se uma nota já foi baixada"""
        return hash_nota in self.cache
    
    def registrar_download(self, hash_nota: str):
        """Registra um download no cache"""
        self.cache[hash_nota] = datetime.now().isoformat()
        self._salvar_cache()
    
    def limpar_cache_antigo(self, dias: int = 90):
        """Remove entradas antigas do cache"""
        limite = datetime.now() - timedelta(days=dias)
        chaves_antigas = []
        
        for hash_nota, data_str in self.cache.items():
            try:
                data = datetime.fromisoformat(data_str)
                if data < limite:
                    chaves_antigas.append(hash_nota)
            except:
                chaves_antigas.append(hash_nota)
        
        for chave in chaves_antigas:
            del self.cache[chave]
        
        if chaves_antigas:
            self._salvar_cache()
            logger.info(f"Removidas {len(chaves_antigas)} entradas antigas do cache")


# ============================================================================
# VALIDADORES
# ============================================================================

class Validadores:
    """Classe com métodos de validação"""
    
    @staticmethod
    def validar_cnpj(cnpj: str) -> bool:
        """Valida formato básico de CNPJ"""
        cnpj_numeros = ''.join(filter(str.isdigit, cnpj))
        return len(cnpj_numeros) == 14
    
    @staticmethod
    def validar_competencia(competencia: str) -> tuple:
        """
        Valida e retorna (mes, ano) da competência
        Retorna (None, None) se inválido
        """
        try:
            partes = competencia.strip().split('/')
            if len(partes) != 2:
                return (None, None)
            
            mes = int(partes[0])
            ano = int(partes[1])
            
            if not (1 <= mes <= 12):
                return (None, None)
            if not (2000 <= ano <= 2100):
                return (None, None)
            
            return (mes, ano)
        except:
            return (None, None)
    
    @staticmethod
    def formatar_cnpj(cnpj: str) -> str:
        """Formata CNPJ para exibição"""
        numeros = ''.join(filter(str.isdigit, cnpj))
        if len(numeros) == 14:
            return f"{numeros[:2]}.{numeros[2:5]}.{numeros[5:8]}/{numeros[8:12]}-{numeros[12:]}"
        return cnpj
    
    @staticmethod
    def limpar_nome_arquivo(nome: str) -> str:
        """Remove caracteres inválidos de nomes de arquivo"""
        invalidos = '<>:"/\\|?*'
        for char in invalidos:
            nome = nome.replace(char, '')
        # Remove espaços extras e limita tamanho
        nome = ' '.join(nome.split())
        return nome[:200].strip()


# ============================================================================
# GERENCIADOR DE PERÍODOS
# ============================================================================

class GerenciadorPeriodos:
    """Gera períodos de consulta baseado em competência"""
    
    @staticmethod
    def gerar_periodos(competencia: str = None) -> list:
        """
        Gera lista de períodos (data_inicio, data_fim) para consulta
        
        Se competência fornecida: retorna período do mês + períodos subsequentes até hoje
        Se não: retorna período vazio (últimos 30 dias do sistema)
        """
        if not competencia:
            return [("", "")]
        
        mes, ano = Validadores.validar_competencia(competencia)
        if mes is None:
            logger.warning(f"Competência inválida: {competencia}")
            return [("", "")]
        
        periodos = []
        
        # Período da competência
        ultimo_dia_mes = calendar.monthrange(ano, mes)[1]
        dt_ini = datetime(ano, mes, 1)
        dt_fim = datetime(ano, mes, ultimo_dia_mes)
        periodos.append((dt_ini.strftime("%d/%m/%Y"), dt_fim.strftime("%d/%m/%Y")))
        
        # Períodos subsequentes até hoje (blocos de 30 dias)
        hoje = datetime.now()
        cursor = dt_fim + timedelta(days=1)
        
        while cursor <= hoje:
            fim_bloco = cursor + timedelta(days=29)
            if fim_bloco > hoje:
                fim_bloco = hoje
            periodos.append((cursor.strftime("%d/%m/%Y"), fim_bloco.strftime("%d/%m/%Y")))
            cursor = fim_bloco + timedelta(days=1)
        
        logger.info(f"Gerados {len(periodos)} períodos de consulta para {competencia}")
        return periodos


# ============================================================================

# ============================================================================
# PROCESSAMENTO DE XML E GERAÇÃO DE RELATÓRIOS
# ============================================================================

class ProcessadorXML:
    """Processa XMLs de NFS-e e extrai dados"""
    
    # Namespaces do XML
    NS = {
        'nfse': 'http://www.sped.fazenda.gov.br/nfse'
    }
    
    @staticmethod
    def extrair_texto(elemento, caminho, default=''):
        """Extrai texto de um elemento XML com tratamento de erro"""
        try:
            elem = elemento.find(caminho, ProcessadorXML.NS)
            return elem.text if elem is not None and elem.text else default
        except:
            return default
    
    @staticmethod
    def extrair_decimal(elemento, caminho, default=0.0):
        """Extrai valor decimal de um elemento XML"""
        try:
            texto = ProcessadorXML.extrair_texto(elemento, caminho, '0')
            return float(texto.replace(',', '.'))
        except:
            return default
    
    @staticmethod
    def processar_xml_emitida(caminho_xml):
        """
        Processa um XML de NFS-e EMITIDA e retorna dicionário com dados
        """
        try:
            tree = ET.parse(caminho_xml)
            root = tree.getroot()
            
            # Encontrar elemento infNFSe
            inf_nfse = root.find('.//nfse:infNFSe', ProcessadorXML.NS)
            if inf_nfse is None:
                logger.warning(f"Elemento infNFSe não encontrado em {caminho_xml}")
                return None
            
            # Encontrar elemento DPS
            dps = inf_nfse.find('.//nfse:DPS/nfse:infDPS', ProcessadorXML.NS)
            
            # Extrair dados
            dados = {
                # Identificação
                'numero_nfse': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:nNFSe'),
                'serie': ProcessadorXML.extrair_texto(dps, './/nfse:serie') if dps is not None else '',
                'data_emissao': ProcessadorXML.extrair_texto(dps, './/nfse:dhEmi') if dps is not None else '',
                'competencia': ProcessadorXML.extrair_texto(dps, './/nfse:dCompet') if dps is not None else '',
                'status': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:cStat'),
                
                # Tomador (quem recebeu o serviço)
                'tomador_cnpj': ProcessadorXML.extrair_texto(dps, './/nfse:toma/nfse:CNPJ') if dps is not None else '',
                'tomador_cpf': ProcessadorXML.extrair_texto(dps, './/nfse:toma/nfse:CPF') if dps is not None else '',
                'tomador_nome': ProcessadorXML.extrair_texto(dps, './/nfse:toma/nfse:xNome') if dps is not None else '',
                'tomador_municipio': ProcessadorXML.extrair_texto(dps, './/nfse:toma/nfse:end/nfse:endNac/nfse:cMun') if dps is not None else '',
                
                # Serviço
                'codigo_servico': ProcessadorXML.extrair_texto(dps, './/nfse:cServ/nfse:cTribNac') if dps is not None else '',
                'descricao_servico': ProcessadorXML.extrair_texto(dps, './/nfse:cServ/nfse:xDescServ') if dps is not None else '',
                
                # Valores
                'valor_servico': ProcessadorXML.extrair_decimal(dps, './/nfse:vServPrest/nfse:vServ') if dps is not None else 0.0,
                'base_calculo': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vBC'),
                'aliquota': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:pAliqAplic'),
                'valor_issqn': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vISSQN'),
                'valor_retido': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vTotalRet'),
                'valor_liquido': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vLiq'),
                'percentual_tributos_sn': ProcessadorXML.extrair_decimal(dps, './/nfse:totTrib/nfse:pTotTribSN') if dps is not None else 0.0,
                
                # Local
                'local_prestacao': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:xLocPrestacao'),
                
                # Arquivo
                'arquivo_xml': os.path.basename(caminho_xml)
            }
            
            # Formatar CNPJ/CPF do tomador
            if dados['tomador_cnpj']:
                cnpj = dados['tomador_cnpj']
                dados['tomador_doc'] = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            elif dados['tomador_cpf']:
                cpf = dados['tomador_cpf']
                dados['tomador_doc'] = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
            else:
                dados['tomador_doc'] = ''
            
            return dados
            
        except Exception as e:
            logger.error(f"Erro ao processar XML {caminho_xml}: {e}")
            return None
    
    @staticmethod
    def processar_xml_recebida(caminho_xml):
        """
        Processa um XML de NFS-e RECEBIDA e retorna dicionário com dados
        """
        try:
            tree = ET.parse(caminho_xml)
            root = tree.getroot()
            
            # Encontrar elemento infNFSe
            inf_nfse = root.find('.//nfse:infNFSe', ProcessadorXML.NS)
            if inf_nfse is None:
                logger.warning(f"Elemento infNFSe não encontrado em {caminho_xml}")
                return None
            
            # Encontrar elemento DPS
            dps = inf_nfse.find('.//nfse:DPS/nfse:infDPS', ProcessadorXML.NS)
            
            # Extrair dados
            dados = {
                # Identificação
                'numero_nfse': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:nNFSe'),
                'serie': ProcessadorXML.extrair_texto(dps, './/nfse:serie') if dps is not None else '',
                'data_emissao': ProcessadorXML.extrair_texto(dps, './/nfse:dhEmi') if dps is not None else '',
                'competencia': ProcessadorXML.extrair_texto(dps, './/nfse:dCompet') if dps is not None else '',
                'status': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:cStat'),
                
                # Prestador (quem emitiu a nota - emissor do serviço)
                'prestador_cnpj': ProcessadorXML.extrair_texto(dps, './/nfse:prest/nfse:CNPJ') if dps is not None else '',
                'prestador_cpf': ProcessadorXML.extrair_texto(dps, './/nfse:prest/nfse:CPF') if dps is not None else '',
                'prestador_nome': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:emit/nfse:xNome'),
                'prestador_im': ProcessadorXML.extrair_texto(dps, './/nfse:prest/nfse:IM') if dps is not None else '',
                'prestador_municipio': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:emit/nfse:enderNac/nfse:cMun'),
                
                # Serviço
                'codigo_servico': ProcessadorXML.extrair_texto(dps, './/nfse:cServ/nfse:cTribNac') if dps is not None else '',
                'descricao_servico': ProcessadorXML.extrair_texto(dps, './/nfse:cServ/nfse:xDescServ') if dps is not None else '',
                
                # Valores
                'valor_servico': ProcessadorXML.extrair_decimal(dps, './/nfse:vServPrest/nfse:vServ') if dps is not None else 0.0,
                'base_calculo': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vBC'),
                'aliquota': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:pAliqAplic'),
                'valor_issqn': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vISSQN'),
                'valor_retido': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vTotalRet'),
                'valor_liquido': ProcessadorXML.extrair_decimal(inf_nfse, './/nfse:valores/nfse:vLiq'),
                'percentual_tributos_sn': ProcessadorXML.extrair_decimal(dps, './/nfse:totTrib/nfse:pTotTribSN') if dps is not None else 0.0,
                
                # Local
                'local_prestacao': ProcessadorXML.extrair_texto(inf_nfse, './/nfse:xLocPrestacao'),
                
                # Arquivo
                'arquivo_xml': os.path.basename(caminho_xml)
            }
            
            # Formatar CNPJ/CPF do prestador
            if dados['prestador_cnpj']:
                cnpj = dados['prestador_cnpj']
                dados['prestador_doc'] = f"{cnpj[:2]}.{cnpj[2:5]}.{cnpj[5:8]}/{cnpj[8:12]}-{cnpj[12:]}"
            elif dados['prestador_cpf']:
                cpf = dados['prestador_cpf']
                dados['prestador_doc'] = f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
            else:
                dados['prestador_doc'] = ''
            
            return dados
            
        except Exception as e:
            logger.error(f"Erro ao processar XML {caminho_xml}: {e}")
            return None


class GeradorRelatorioExcel:
    """Gera relatórios Excel a partir de XMLs de NFS-e"""
    
    @staticmethod
    def aplicar_estilo_cabecalho(ws, linha=1):
        """Aplica estilo ao cabeçalho da planilha"""
        # Cor de fundo azul
        fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        # Fonte branca e negrito
        font = Font(color="FFFFFF", bold=True, size=11)
        # Alinhamento centralizado
        alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # Borda
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[linha]:
            cell.fill = fill
            cell.font = font
            cell.alignment = alignment
            cell.border = border
    
    @staticmethod
    def ajustar_largura_colunas(ws):
        """Ajusta largura das colunas automaticamente"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Definir largura (mínimo 10, máximo 50)
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def gerar_relatorio_emitidas(lista_xmls, caminho_saida):
        """
        Gera relatório Excel de NFS-e EMITIDAS
        
        Args:
            lista_xmls: Lista de caminhos completos dos arquivos XML
            caminho_saida: Caminho completo do arquivo Excel de saída
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "NFS-e Emitidas"
            
            # Cabeçalhos
            headers = [
                "Número NFS-e", "Série", "Data Emissão", "Competência", "Status",
                "Tomador (CNPJ/CPF)", "Tomador (Nome)", "Município Tomador",
                "Código Serviço", "Descrição Serviço",
                "Valor Serviço (R$)", "Base Cálculo (R$)", "Alíquota (%)",
                "Valor ISSQN (R$)", "Valor Retido (R$)", "Valor Líquido (R$)",
                "% Tributos SN", "Local Prestação", "Arquivo XML"
            ]
            
            ws.append(headers)
            GeradorRelatorioExcel.aplicar_estilo_cabecalho(ws)
            
            # Processar cada XML
            total_valor_servico = 0.0
            total_issqn = 0.0
            total_liquido = 0.0
            
            for xml_path in lista_xmls:
                dados = ProcessadorXML.processar_xml_emitida(xml_path)
                if dados:
                    linha = [
                        dados['numero_nfse'],
                        dados['serie'],
                        dados['data_emissao'][:10] if dados['data_emissao'] else '',  # Apenas data
                        dados['competencia'],
                        dados['status'],
                        dados['tomador_doc'],
                        dados['tomador_nome'],
                        dados['tomador_municipio'],
                        dados['codigo_servico'],
                        dados['descricao_servico'],
                        dados['valor_servico'],
                        dados['base_calculo'],
                        dados['aliquota'],
                        dados['valor_issqn'],
                        dados['valor_retido'],
                        dados['valor_liquido'],
                        dados['percentual_tributos_sn'],
                        dados['local_prestacao'],
                        dados['arquivo_xml']
                    ]
                    ws.append(linha)
                    
                    # Acumular totais
                    total_valor_servico += dados['valor_servico']
                    total_issqn += dados['valor_issqn']
                    total_liquido += dados['valor_liquido']
            
            # Adicionar linha de totais
            if len(lista_xmls) > 0:
                ws.append([])  # Linha em branco
                linha_total = ws.max_row + 1
                ws.append([
                    "", "", "", "", "",
                    "", "", "", "",
                    "TOTAIS:",
                    total_valor_servico,
                    "",
                    "",
                    total_issqn,
                    "",
                    total_liquido,
                    "", "", ""
                ])
                
                # Estilo da linha de totais
                for cell in ws[linha_total]:
                    cell.font = Font(bold=True)
                    if cell.column >= 11 and cell.column <= 16:  # Colunas de valores
                        cell.number_format = '#,##0.00'
            
            # Formatar colunas de valores
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row-2, min_col=11, max_col=17):
                for cell in row:
                    if cell.column <= 16:  # Colunas monetárias
                        cell.number_format = '#,##0.00'
                    else:  # Coluna de percentual
                        cell.number_format = '0.00'
            
            # Ajustar larguras
            GeradorRelatorioExcel.ajustar_largura_colunas(ws)
            
            # Salvar
            wb.save(caminho_saida)
            logger.info(f"Relatório de emitidas salvo: {caminho_saida}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de emitidas: {e}")
            return False
    
    @staticmethod
    def gerar_relatorio_recebidas(lista_xmls, caminho_saida):
        """
        Gera relatório Excel de NFS-e RECEBIDAS
        
        Args:
            lista_xmls: Lista de caminhos completos dos arquivos XML
            caminho_saida: Caminho completo do arquivo Excel de saída
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "NFS-e Recebidas"
            
            # Cabeçalhos
            headers = [
                "Número NFS-e", "Série", "Data Emissão", "Competência", "Status",
                "Prestador (CNPJ/CPF)", "Prestador (Nome)", "Prestador (IM)", "Município Prestador",
                "Código Serviço", "Descrição Serviço",
                "Valor Serviço (R$)", "Base Cálculo (R$)", "Alíquota (%)",
                "Valor ISSQN (R$)", "Valor Retido (R$)", "Valor Líquido (R$)",
                "% Tributos SN", "Local Prestação", "Arquivo XML"
            ]
            
            ws.append(headers)
            GeradorRelatorioExcel.aplicar_estilo_cabecalho(ws)
            
            # Processar cada XML
            total_valor_servico = 0.0
            total_issqn = 0.0
            total_liquido = 0.0
            
            for xml_path in lista_xmls:
                dados = ProcessadorXML.processar_xml_recebida(xml_path)
                if dados:
                    linha = [
                        dados['numero_nfse'],
                        dados['serie'],
                        dados['data_emissao'][:10] if dados['data_emissao'] else '',
                        dados['competencia'],
                        dados['status'],
                        dados['prestador_doc'],
                        dados['prestador_nome'],
                        dados['prestador_im'],
                        dados['prestador_municipio'],
                        dados['codigo_servico'],
                        dados['descricao_servico'],
                        dados['valor_servico'],
                        dados['base_calculo'],
                        dados['aliquota'],
                        dados['valor_issqn'],
                        dados['valor_retido'],
                        dados['valor_liquido'],
                        dados['percentual_tributos_sn'],
                        dados['local_prestacao'],
                        dados['arquivo_xml']
                    ]
                    ws.append(linha)
                    
                    # Acumular totais
                    total_valor_servico += dados['valor_servico']
                    total_issqn += dados['valor_issqn']
                    total_liquido += dados['valor_liquido']
            
            # Adicionar linha de totais
            if len(lista_xmls) > 0:
                ws.append([])
                linha_total = ws.max_row + 1
                ws.append([
                    "", "", "", "", "",
                    "", "", "", "",
                    "TOTAIS:",
                    total_valor_servico,
                    "",
                    "",
                    total_issqn,
                    "",
                    total_liquido,
                    "", "", ""
                ])
                
                # Estilo da linha de totais
                for cell in ws[linha_total]:
                    cell.font = Font(bold=True)
                    if cell.column >= 12 and cell.column <= 17:
                        cell.number_format = '#,##0.00'
            
            # Formatar colunas de valores
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row-2, min_col=12, max_col=18):
                for cell in row:
                    if cell.column <= 17:
                        cell.number_format = '#,##0.00'
                    else:
                        cell.number_format = '0.00'
            
            # Ajustar larguras
            GeradorRelatorioExcel.ajustar_largura_colunas(ws)
            
            # Salvar
            wb.save(caminho_saida)
            logger.info(f"Relatório de recebidas salvo: {caminho_saida}")
            return True
            
        except Exception as e:
            logger.error(f"Erro ao gerar relatório de recebidas: {e}")
            return False
    
    @staticmethod
    def gerar_relatorios_automaticos(caminho_empresa, competencia):
        """
        Gera relatórios automáticos de EMITIDAS e RECEBIDAS para uma competência
        
        Args:
            caminho_empresa: Caminho da pasta da empresa
            competencia: Competência no formato MM-AAAA
        
        Returns:
            Tuple (sucesso_emitidas, sucesso_recebidas, caminho_relatorio_emitidas, caminho_relatorio_recebidas)
        """
        sucesso_emitidas = False
        sucesso_recebidas = False
        caminho_rel_emitidas = None
        caminho_rel_recebidas = None
        
        # Processar EMITIDAS
        pasta_emitidas = os.path.join(caminho_empresa, "EMITIDAS", competencia)
        if os.path.exists(pasta_emitidas):
            xmls_emitidas = []
            for root, dirs, files in os.walk(pasta_emitidas):
                if 'XML' in root:
                    for file in files:
                        if file.lower().endswith('.xml'):
                            xmls_emitidas.append(os.path.join(root, file))
            
            if xmls_emitidas:
                caminho_rel_emitidas = os.path.join(
                    caminho_empresa, 
                    "EMITIDAS",
                    f"Relatório_Emitidas_{competencia}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                sucesso_emitidas = GeradorRelatorioExcel.gerar_relatorio_emitidas(
                    xmls_emitidas,
                    caminho_rel_emitidas
                )
        
        # Processar RECEBIDAS
        pasta_recebidas = os.path.join(caminho_empresa, "RECEBIDAS", competencia)
        if os.path.exists(pasta_recebidas):
            xmls_recebidas = []
            for root, dirs, files in os.walk(pasta_recebidas):
                if 'XML' in root:
                    for file in files:
                        if file.lower().endswith('.xml'):
                            xmls_recebidas.append(os.path.join(root, file))
            
            if xmls_recebidas:
                caminho_rel_recebidas = os.path.join(
                    caminho_empresa,
                    "RECEBIDAS",
                    f"Relatório_Recebidas_{competencia}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                sucesso_recebidas = GeradorRelatorioExcel.gerar_relatorio_recebidas(
                    xmls_recebidas,
                    caminho_rel_recebidas
                )
        
        return sucesso_emitidas, sucesso_recebidas, caminho_rel_emitidas, caminho_rel_recebidas


# APLICAÇÃO PRINCIPAL
# ============================================================================

class NFSeDownloaderApp:
    """Aplicação principal com interface gráfica"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("NFSe Downloader Pro - V18 (Interface com Abas)")
        self.root.geometry("800x900")
        
        # Gerenciadores
        self.crypto = CryptoManager()
        self.cache = DownloadCache()
        self.validador = Validadores()
        
        # Variáveis
        self.path_download = tk.StringVar()
        self.competencia_filtro = tk.StringVar()
        self.tipo_download = tk.StringVar(value="ambos")
        self.var_usa_certificado = tk.BooleanVar(value=False)
        self.var_usar_cache = tk.BooleanVar(value=True)
        self.var_modo_headless = tk.BooleanVar(value=False)
        self.var_baixar_emitidas = tk.BooleanVar(value=True)
        self.var_baixar_recebidas = tk.BooleanVar(value=False)
        
        self.empresas = []
        self.cnpj_em_edicao = None
        self.thread_ativa = None
        self.cancelar_flag = False
        
        # Carregar dados
        self.carregar_dados()
        
        # Criar interface
        self._criar_interface()
        
        # Preencher lista de empresas
        self.atualizar_lista_visual()
        
        # Limpar cache antigo ao iniciar
        self.cache.limpar_cache_antigo()
        
        logger.info("Aplicação iniciada - V16 Melhorado")
    
    def _criar_interface(self):
        """Cria interface com sistema de abas"""
        
        # ===== NOTEBOOK (SISTEMA DE ABAS) =====
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=5, pady=5)
        
        # ===== ABA 1: EMPRESAS =====
        self.tab_empresas = tk.Frame(self.notebook)
        self.notebook.add(self.tab_empresas, text="📁 Empresas")
        self._criar_aba_empresas()
        
        # ===== ABA 2: CONFIGURAÇÃO =====
        self.tab_configuracao = tk.Frame(self.notebook)
        self.notebook.add(self.tab_configuracao, text="⚙️ Configuração")
        self._criar_aba_configuracao()
        
        # ===== ABA 3: RELATÓRIOS =====
        self.tab_relatorios = tk.Frame(self.notebook)
        self.notebook.add(self.tab_relatorios, text="📊 Relatórios")
        self._criar_aba_relatorios()
        
        # Bind para atualizar preview ao mudar de aba
        self.notebook.bind('<<NotebookTabChanged>>', self._atualizar_preview_config)
    
    def _criar_aba_empresas(self):
        """Cria conteúdo da aba Empresas"""
        
        # ===== CADASTRO DE EMPRESA =====
        self.frame_cadastro = tk.LabelFrame(
            self.tab_empresas,
            text="📝 Cadastrar / Editar Empresa",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        self.frame_cadastro.pack(fill="x", padx=10, pady=10)
        
        # Nome
        tk.Label(self.frame_cadastro, text="Nome da Empresa:").grid(row=0, column=0, sticky="w", pady=2)
        self.entry_nome = tk.Entry(self.frame_cadastro, width=40)
        self.entry_nome.grid(row=0, column=1, columnspan=2, padx=5, pady=2, sticky="w")
        
        # CNPJ
        tk.Label(self.frame_cadastro, text="CNPJ:").grid(row=1, column=0, sticky="w", pady=2)
        self.entry_cnpj = tk.Entry(self.frame_cadastro, width=20)
        self.entry_cnpj.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        self.entry_cnpj.bind('<FocusOut>', self._validar_cnpj_campo)
        
        # Senha Portal (ao lado do CNPJ)
        self.lbl_senha_portal = tk.Label(self.frame_cadastro, text="Senha Portal:")
        self.lbl_senha_portal.grid(row=1, column=2, sticky="w", padx=(20, 0), pady=2)
        self.entry_senha_portal = tk.Entry(self.frame_cadastro, width=20, show="*")
        self.entry_senha_portal.grid(row=1, column=3, padx=5, pady=2, sticky="w")
        
        # Checkbox certificado
        self.chk_certificado = tk.Checkbutton(
            self.frame_cadastro,
            text="🔒 Acessar com Certificado Digital (Arquivo .PFX)",
            variable=self.var_usa_certificado,
            command=self.toggle_campos_login
        )
        self.chk_certificado.grid(row=2, column=1, columnspan=3, sticky="w", pady=5)
        
        # Campos PFX (em linha separada, ocultos por padrão)
        self.frame_pfx = tk.Frame(self.frame_cadastro)
        self.frame_pfx.grid(row=3, column=0, columnspan=4, sticky="ew", pady=5)
        
        tk.Label(self.frame_pfx, text="Arquivo .PFX:").pack(side="left", padx=(0, 5))
        self.entry_pfx = tk.Entry(self.frame_pfx, width=35)
        self.entry_pfx.pack(side="left", padx=5)
        self.btn_pfx = tk.Button(self.frame_pfx, text="📁", command=self.buscar_pfx, width=3)
        self.btn_pfx.pack(side="left", padx=5)
        
        tk.Label(self.frame_pfx, text="Senha PFX:").pack(side="left", padx=(10, 5))
        self.entry_senha_pfx = tk.Entry(self.frame_pfx, width=15, show="*")
        self.entry_senha_pfx.pack(side="left", padx=5)
        
        # Botões de ação
        frame_botoes_cadastro = tk.Frame(self.frame_cadastro)
        frame_botoes_cadastro.grid(row=4, column=0, columnspan=4, pady=10)
        
        self.btn_salvar = tk.Button(
            frame_botoes_cadastro,
            text="💾 Salvar Nova",
            font=("Arial", 9, "bold"),
            bg="#4CAF50",
            fg="white",
            command=self.salvar_empresa_action,
            cursor="hand2",
            width=15
        )
        self.btn_salvar.pack(side="left", padx=5)
        
        self.btn_cancelar = tk.Button(
            frame_botoes_cadastro,
            text="❌ Cancelar",
            font=("Arial", 9),
            command=self.limpar_campos,
            cursor="hand2",
            width=12,
            state="disabled"
        )
        self.btn_cancelar.pack(side="left", padx=5)
        
        # ===== LISTA DE EMPRESAS =====
        frame_lista = tk.LabelFrame(
            self.tab_empresas,
            text="📋 Empresas Cadastradas",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_lista.pack(fill="both", expand=True, padx=10, pady=10)
        
        # TreeView com seleção múltipla
        columns = ('nome', 'cnpj', 'tipo_acesso')
        self.tree = ttk.Treeview(frame_lista, columns=columns, show='headings', height=8, selectmode='extended')
        self.tree.heading('nome', text='Nome da Empresa')
        self.tree.heading('cnpj', text='CNPJ')
        self.tree.heading('tipo_acesso', text='Tipo Acesso')
        self.tree.column('nome', width=300)
        self.tree.column('cnpj', width=150)
        self.tree.column('tipo_acesso', width=120, anchor="center")
        self.tree.pack(side="left", fill="both", expand=True)
        
        # Scrollbar
        scrollbar = ttk.Scrollbar(frame_lista, orient="vertical", command=self.tree.yview)
        scrollbar.pack(side="right", fill="y")
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        # Botões da lista
        frame_botoes_lista = tk.Frame(self.tab_empresas)
        frame_botoes_lista.pack(pady=5)
        
        tk.Button(
            frame_botoes_lista,
            text="📝 Editar Selecionada",
            command=self.carregar_para_edicao,
            cursor="hand2"
        ).pack(side="left", padx=10)
        
        tk.Button(
            frame_botoes_lista,
            text="🗑️ Remover Selecionada",
            command=self.remover_empresa,
            fg="red",
            cursor="hand2"
        ).pack(side="left", padx=10)
        
        # Separador visual
        ttk.Separator(frame_botoes_lista, orient="vertical").pack(side="left", fill="y", padx=10)
        
        # Botões de seleção múltipla
        tk.Button(
            frame_botoes_lista,
            text="☑️ Selecionar Todas",
            command=self.selecionar_todas_empresas,
            bg="#E3F2FD",
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        tk.Button(
            frame_botoes_lista,
            text="⬜ Desmarcar Todas",
            command=self.desmarcar_todas_empresas,
            bg="#FFF9C4",
            cursor="hand2"
        ).pack(side="left", padx=5)
        
        # Label contador
        self.lbl_selecionadas = tk.Label(
            frame_botoes_lista,
            text="💡 Use CTRL+Clique para múltiplas | SHIFT+Clique para faixa",
            font=("Arial", 8),
            fg="gray"
        )
        self.lbl_selecionadas.pack(side="left", padx=10)
        
        # Binding para atualizar contador
        self.tree.bind('<<TreeviewSelect>>', self.atualizar_contador_selecao)
        
        # Botão para ir para configuração
        frame_navegacao = tk.Frame(self.tab_empresas, bg="#f0f0f0", relief="raised", bd=1)
        frame_navegacao.pack(fill="x", side="bottom", pady=0)
        
        tk.Frame(frame_navegacao, height=10, bg="#f0f0f0").pack()
        
        tk.Button(
            frame_navegacao,
            text="Próximo: Configurar Download →",
            font=("Arial", 10, "bold"),
            bg="#2196F3",
            fg="white",
            command=lambda: self.notebook.select(1),  # Muda para aba 2
            cursor="hand2",
            width=30,
            height=2
        ).pack(pady=10)
    
    def _criar_aba_configuracao(self):
        """Cria conteúdo da aba Configuração"""
        
        # Frame principal com scroll
        canvas = tk.Canvas(self.tab_configuracao)
        scrollbar = ttk.Scrollbar(self.tab_configuracao, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # ===== PASTA DE DOWNLOAD =====
        frame_pasta = tk.LabelFrame(
            scrollable_frame,
            text="📁 Pasta Raiz de Download",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_pasta.pack(fill="x", padx=10, pady=10)
        
        frame_pasta_input = tk.Frame(frame_pasta)
        frame_pasta_input.pack(fill="x")
        
        entry_pasta = tk.Entry(frame_pasta_input, textvariable=self.path_download, width=60)
        entry_pasta.pack(side="left", padx=(0, 5))
        
        tk.Button(
            frame_pasta_input,
            text="📁 Alterar Pasta...",
            command=self.selecionar_pasta,
            cursor="hand2"
        ).pack(side="left")
        
        # ===== COMPETÊNCIA =====
        frame_comp = tk.LabelFrame(
            scrollable_frame,
            text="📅 Filtrar por Competência",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_comp.pack(fill="x", padx=10, pady=10)
        
        frame_comp_input = tk.Frame(frame_comp)
        frame_comp_input.pack()
        
        tk.Entry(frame_comp_input, textvariable=self.competencia_filtro, width=15).pack(side="left", padx=5)
        tk.Label(frame_comp_input, text="(Ex: 01/2026) - Deixe vazio para últimos 30 dias", fg="gray").pack(side="left")
        
        # ===== TIPO DE DOWNLOAD =====
        frame_tipo = tk.LabelFrame(
            scrollable_frame,
            text="📥 O que baixar?",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_tipo.pack(fill="x", padx=10, pady=10)
        
        # Frame interno para organizar horizontalmente
        frame_tipo_opcoes = tk.Frame(frame_tipo)
        frame_tipo_opcoes.pack(anchor="w")
        
        tk.Radiobutton(
            frame_tipo_opcoes,
            text="Apenas XML",
            variable=self.tipo_download,
            value="xml",
            command=self._atualizar_preview_config
        ).pack(side="left", padx=(0, 20))
        
        tk.Radiobutton(
            frame_tipo_opcoes,
            text="Apenas PDF",
            variable=self.tipo_download,
            value="pdf",
            command=self._atualizar_preview_config
        ).pack(side="left", padx=(0, 20))
        
        tk.Radiobutton(
            frame_tipo_opcoes,
            text="PDF + XML (Ambos)",
            variable=self.tipo_download,
            value="ambos",
            command=self._atualizar_preview_config
        ).pack(side="left")
        
        # ===== TIPO DE NFS-E =====
        frame_nfse = tk.LabelFrame(
            scrollable_frame,
            text="📋 Tipo de NFS-e para baixar",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_nfse.pack(fill="x", padx=10, pady=10)
        
        tk.Checkbutton(
            frame_nfse,
            text="📤 NFS-e EMITIDAS (notas que você emitiu para clientes)",
            variable=self.var_baixar_emitidas,
            command=self._atualizar_preview_config
        ).pack(anchor="w", pady=2)
        
        tk.Checkbutton(
            frame_nfse,
            text="📥 NFS-e RECEBIDAS (notas que você recebeu de prestadores)",
            variable=self.var_baixar_recebidas,
            command=self._atualizar_preview_config
        ).pack(anchor="w", pady=2)
        
        # ===== OPÇÕES AVANÇADAS =====
        frame_opcoes = tk.LabelFrame(
            scrollable_frame,
            text="⚙️ Opções Avançadas",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold")
        )
        frame_opcoes.pack(fill="x", padx=10, pady=10)
        
        tk.Checkbutton(
            frame_opcoes,
            text="⚡ Usar cache (evitar downloads duplicados)",
            variable=self.var_usar_cache
        ).pack(anchor="w", pady=2)
        
        tk.Checkbutton(
            frame_opcoes,
            text="👁️ Modo headless (navegador oculto)",
            variable=self.var_modo_headless
        ).pack(anchor="w", pady=2)
        
        # ===== PREVIEW DAS CONFIGURAÇÕES =====
        frame_preview = tk.LabelFrame(
            scrollable_frame,
            text="👁️ Preview das Configurações",
            padx=10,
            pady=10,
            font=("Arial", 10, "bold"),
            bg="#E8F5E9"
        )
        frame_preview.pack(fill="x", padx=10, pady=10)
        
        self.lbl_preview = tk.Label(
            frame_preview,
            text="",
            justify="left",
            font=("Consolas", 9),
            bg="#E8F5E9"
        )
        self.lbl_preview.pack(anchor="w")
        
        # Atualizar preview inicial
        self._atualizar_preview_config()
        
        # ===== BOTÕES DE AÇÃO =====
        frame_acoes = tk.Frame(scrollable_frame, bg="#f0f0f0", relief="raised", bd=1)
        frame_acoes.pack(fill="x", pady=10)
        
        tk.Frame(frame_acoes, height=10, bg="#f0f0f0").pack()
        
        frame_botoes = tk.Frame(frame_acoes, bg="#f0f0f0")
        frame_botoes.pack(pady=10)
        
        self.btn_run = tk.Button(
            frame_botoes,
            text="🚀 INICIAR DOWNLOADS",
            font=("Arial", 12, "bold"),
            bg="#4CAF50",
            fg="white",
            command=self.iniciar_thread,
            cursor="hand2",
            width=25,
            height=2
        )
        self.btn_run.pack(side="left", padx=10)
        
        self.btn_cancelar_exec = tk.Button(
            frame_botoes,
            text="⛔ CANCELAR",
            font=("Arial", 12, "bold"),
            bg="#f44336",
            fg="white",
            command=self.cancelar_execucao,
            cursor="hand2",
            width=15,
            height=2,
            state="disabled"
        )
        self.btn_cancelar_exec.pack(side="left", padx=10)
        
        # ===== STATUS E LOG =====
        self.lbl_status = tk.Label(
            scrollable_frame,
            text="⏳ Aguardando início...",
            fg="blue",
            font=("Arial", 10)
        )
        self.lbl_status.pack(pady=5)
        
        frame_log = tk.LabelFrame(scrollable_frame, text="Log de Execução", padx=5, pady=5)
        frame_log.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        
        self.text_log = scrolledtext.ScrolledText(
            frame_log,
            height=12,
            width=80,
            state='disabled',
            font=("Consolas", 8)
        )
        self.text_log.pack(fill="both", expand=True)
        
        # Tags para cores no log
        self.text_log.tag_config("INFO", foreground="black")
        self.text_log.tag_config("SUCCESS", foreground="green", font=("Consolas", 8, "bold"))
        self.text_log.tag_config("WARNING", foreground="orange")
        self.text_log.tag_config("ERROR", foreground="red", font=("Consolas", 8, "bold"))
    
    def _criar_aba_relatorios(self):
        """Cria conteúdo da aba Relatórios"""
        
        # Título
        tk.Label(
            self.tab_relatorios,
            text="📊 Gerar Relatórios Excel a partir de XMLs",
            font=("Arial", 14, "bold")
        ).pack(pady=20)
        
        # Descrição
        frame_desc = tk.Frame(self.tab_relatorios)
        frame_desc.pack(padx=20, pady=10)
        
        tk.Label(
            frame_desc,
            text="Selecione arquivos XML de NFS-e e gere relatórios Excel formatados\n"
                 "Os XMLs podem ser de qualquer local (Downloads, Email, Backup, etc.)",
            justify="center",
            fg="gray"
        ).pack()
        
        # Botão principal
        tk.Button(
            self.tab_relatorios,
            text="📊 GERAR RELATÓRIOS EXCEL",
            font=("Arial", 12, "bold"),
            bg="#2196F3",
            fg="white",
            command=self.gerar_relatorios_manual,
            cursor="hand2",
            width=30,
            height=3
        ).pack(pady=20)
        
        # Informações adicionais
        frame_info = tk.LabelFrame(
            self.tab_relatorios,
            text="ℹ️ Como Funciona",
            padx=20,
            pady=15
        )
        frame_info.pack(fill="x", padx=20, pady=10)
        
        info_text = """
1. Clique no botão acima para abrir a janela de geração de relatórios

2. Selecione os arquivos XML das notas fiscais
   • Podem ser de qualquer pasta do computador
   • Selecione múltiplos arquivos de uma vez (CTRL + Clique)

3. Escolha o tipo de relatório:
   • Emitidas: Para notas que você emitiu
   • Recebidas: Para notas que você recebeu

4. Escolha onde salvar o arquivo Excel

5. O sistema processará todos os XMLs e gerará um relatório profissional
   com todas as informações organizadas e totalizadas
"""
        
        tk.Label(
            frame_info,
            text=info_text,
            justify="left",
            font=("Arial", 9)
        ).pack(anchor="w")
        
        # Exemplo visual
        frame_exemplo = tk.LabelFrame(
            self.tab_relatorios,
            text="📋 O Relatório Contém",
            padx=20,
            pady=15
        )
        frame_exemplo.pack(fill="x", padx=20, pady=10)
        
        colunas = [
            "• Número da NFS-e",
            "• Data de Emissão e Competência",
            "• Dados do Tomador/Prestador",
            "• Código e Descrição do Serviço",
            "• Valores (Serviço, Base de Cálculo, ISSQN, Líquido)",
            "• Alíquota e Percentual de Tributos",
            "• Local de Prestação",
            "• Totais calculados automaticamente"
        ]
        
        tk.Label(
            frame_exemplo,
            text="\n".join(colunas),
            justify="left",
            font=("Arial", 9)
        ).pack(anchor="w")
    
    def _atualizar_preview_config(self, event=None):
        """Atualiza o preview das configurações"""
        try:
            qtd_empresas = len(self.tree.selection())
            comp = self.competencia_filtro.get().strip() or "Últimos 30 dias"
            
            tipo_dl = self.tipo_download.get()
            tipo_dl_texto = {
                "xml": "Apenas XML",
                "pdf": "Apenas PDF",
                "ambos": "PDF + XML (Ambos)"
            }.get(tipo_dl, "PDF + XML")
            
            emitidas = "SIM" if self.var_baixar_emitidas.get() else "NÃO"
            recebidas = "SIM" if self.var_baixar_recebidas.get() else "NÃO"
            
            pasta = self.path_download.get() or "(não definida)"
            if len(pasta) > 50:
                pasta = pasta[:47] + "..."
            
            preview = f"""📊 Empresas selecionadas: {qtd_empresas}
📅 Competência: {comp}
📥 Download: {tipo_dl_texto}
📤 Emitidas: {emitidas}  |  📥 Recebidas: {recebidas}
📁 Pasta: {pasta}"""
            
            self.lbl_preview.config(text=preview)
        except:
            pass

        """Valida CNPJ ao sair do campo"""
        cnpj = self.entry_cnpj.get().strip()
        if cnpj and not self.validador.validar_cnpj(cnpj):
            messagebox.showwarning(
                "CNPJ Inválido",
                "O CNPJ digitado não possui formato válido.\n"
                "Deve conter 14 dígitos numéricos."
            )
            self.entry_cnpj.focus()
    
    def _validar_competencia_campo(self, event=None):
        """Valida competência ao sair do campo"""
        comp = self.competencia_filtro.get().strip()
        if comp:
            mes, ano = self.validador.validar_competencia(comp)
            if mes is None:
                messagebox.showwarning(
                    "Competência Inválida",
                    "Formato esperado: MM/AAAA\n"
                    "Exemplo: 01/2026"
                )
                self.entry_competencia.focus()
    
    # ========================================================================
    # MÉTODOS DE INTERFACE
    # ========================================================================
    
    def adicionar_log(self, mensagem: str, nivel: str = "INFO"):
        """Adiciona mensagem ao log visual"""
        def _add():
            self.text_log.config(state='normal')
            timestamp = datetime.now().strftime("%H:%M:%S")
            
            # Cores por nível
            cores = {
                "INFO": "black",
                "WARNING": "orange",
                "ERROR": "red",
                "SUCCESS": "green"
            }
            cor = cores.get(nivel, "black")
            
            # Inserir com tag
            tag = f"tag_{nivel}"
            self.text_log.tag_config(tag, foreground=cor)
            self.text_log.insert('end', f"[{timestamp}] {mensagem}\n", tag)
            self.text_log.see('end')
            self.text_log.config(state='disabled')
        
        self.root.after(0, _add)
    
    def atualizar_status(self, texto: str, cor: str = "blue"):
        """Atualiza label de status de forma thread-safe"""
        def _update():
            self.lbl_status.config(text=texto, fg=cor)
        self.root.after(0, _update)
    
    
    def _validar_cnpj_campo(self, event=None):
        """Valida CNPJ ao sair do campo"""
        cnpj = self.entry_cnpj.get().strip()
        if cnpj and not self.validador.validar_cnpj(cnpj):
            messagebox.showwarning(
                "CNPJ Inválido",
                "O CNPJ informado não é válido.\n"
                "Por favor, verifique e corrija."
            )
            self.entry_cnpj.focus()
    
    def buscar_pfx(self):
        """Abre diálogo para selecionar arquivo PFX"""
        arquivo = filedialog.askopenfilename(
            title="Selecione o Certificado Digital",
            filetypes=[("Arquivo PFX", "*.pfx"), ("Todos", "*.*")]
        )
        if arquivo:
            self.entry_pfx.delete(0, 'end')
            self.entry_pfx.insert(0, arquivo)
    
    def toggle_campos_login(self):
        """Alterna entre campos de senha e certificado"""
        if self.var_usa_certificado.get():
            # Mostrar campos PFX (frame já criado, só precisa mostrar)
            self.lbl_senha_portal.grid_forget()
            self.entry_senha_portal.grid_forget()
            self.frame_pfx.grid(row=3, column=0, columnspan=4, sticky="ew", pady=5)
        else:
            # Mostrar campo senha
            self.frame_pfx.grid_forget()
            self.lbl_senha_portal.grid(row=1, column=2, sticky="w", padx=(20, 0), pady=2)
            self.entry_senha_portal.grid(row=1, column=3, padx=5, pady=2, sticky="w")
    
    def selecionar_pasta(self):
        """Seleciona pasta raiz de downloads"""
        caminho = filedialog.askdirectory(
            title="Selecione a pasta RAIZ para downloads"
        )
        if caminho:
            self.path_download.set(caminho)
            self.salvar_tudo()
            self.adicionar_log(f"Pasta de download alterada: {caminho}", "INFO")
    
    # ========================================================================
    # PERSISTÊNCIA DE DADOS
    # ========================================================================
    
    def carregar_dados(self):
        """Carrega dados do arquivo de configuração"""
        if os.path.exists(ARQUIVO_CONFIG):
            try:
                with open(ARQUIVO_CONFIG, 'r', encoding='utf-8') as f:
                    dados = json.load(f)
                    
                    if isinstance(dados, list):
                        self.empresas = dados
                    else:
                        self.empresas = dados.get('empresas', [])
                        self.path_download.set(dados.get('pasta_padrao', ''))
                    
                    # Descriptografar senhas
                    for emp in self.empresas:
                        if 'senha' in emp and emp['senha']:
                            emp['senha'] = self.crypto.descriptografar(emp['senha'])
                        if 'senha_pfx' in emp and emp['senha_pfx']:
                            emp['senha_pfx'] = self.crypto.descriptografar(emp['senha_pfx'])
                    
                    logger.info(f"Carregadas {len(self.empresas)} empresas")
            except Exception as e:
                logger.error(f"Erro ao carregar dados: {e}")
                self.empresas = []
        else:
            self.path_download.set("C:/Notas Fiscais")
    
    def salvar_tudo(self):
        """Salva todos os dados no arquivo de configuração"""
        try:
            # Criptografar senhas antes de salvar
            empresas_salvar = []
            for emp in self.empresas:
                emp_copy = emp.copy()
                if 'senha' in emp_copy and emp_copy['senha']:
                    emp_copy['senha'] = self.crypto.criptografar(emp_copy['senha'])
                if 'senha_pfx' in emp_copy and emp_copy['senha_pfx']:
                    emp_copy['senha_pfx'] = self.crypto.criptografar(emp_copy['senha_pfx'])
                empresas_salvar.append(emp_copy)
            
            dados_completos = {
                "pasta_padrao": self.path_download.get(),
                "empresas": empresas_salvar
            }
            
            with open(ARQUIVO_CONFIG, 'w', encoding='utf-8') as f:
                json.dump(dados_completos, f, indent=4, ensure_ascii=False)
            
            logger.info("Dados salvos com sucesso")
        except Exception as e:
            logger.error(f"Erro ao salvar dados: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar dados:\n{e}")
    
    # ========================================================================
    # GERENCIAMENTO DE EMPRESAS
    # ========================================================================
    
    def limpar_campos(self):
        """Limpa todos os campos do formulário"""
        self.entry_nome.delete(0, 'end')
        self.entry_cnpj.delete(0, 'end')
        self.entry_senha_portal.delete(0, 'end')
        self.entry_pfx.delete(0, 'end')
        self.entry_senha_pfx.delete(0, 'end')
        self.var_usa_certificado.set(False)
        self.toggle_campos_login()
        
        self.cnpj_em_edicao = None
        self.btn_salvar.config(
            text="💾 Salvar Nova",
            bg="#4CAF50",
            fg="white"
        )
        self.btn_cancelar.config(state="disabled")
        self.frame_cadastro.config(text="Cadastrar Nova Empresa")
    
    def carregar_para_edicao(self):
        """Carrega dados de empresa selecionada para edição"""
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showinfo("Atenção", "Selecione uma empresa na lista")
            return
        
        # Pegar apenas o primeiro item se múltiplos estiverem selecionados
        if len(selected_item) > 1:
            selected_item = selected_item[0]
        
        item = self.tree.item(selected_item)
        cnpj_formatado = str(item['values'][1])
        
        # Remover formatação do CNPJ (pontos, traços, barras)
        cnpj = cnpj_formatado.replace('.', '').replace('/', '').replace('-', '')
        
        empresa_dados = next(
            (e for e in self.empresas if str(e.get('cnpj')) == cnpj),
            None
        )
        
        if not empresa_dados:
            messagebox.showerror("Erro", "Empresa não encontrada nos dados")
            return
        
        self.limpar_campos()
        
        # Preencher campos
        self.entry_nome.insert(0, empresa_dados.get('nome', ''))
        self.entry_cnpj.insert(0, cnpj)
        
        usa_cert = empresa_dados.get('usa_certificado', False)
        self.var_usa_certificado.set(usa_cert)
        
        if usa_cert:
            self.entry_pfx.insert(0, empresa_dados.get('caminho_pfx', ''))
            self.entry_senha_pfx.insert(0, empresa_dados.get('senha_pfx', ''))
        else:
            self.entry_senha_portal.insert(0, empresa_dados.get('senha', ''))
        
        self.toggle_campos_login()
        
        self.cnpj_em_edicao = cnpj
        self.btn_salvar.config(
            text="💾 Salvar Alterações",
            bg="#FFC107",
            fg="black"
        )
        self.btn_cancelar.config(state="normal")
        
        # Mudar para aba de empresas caso esteja em outra
        self.notebook.select(0)
        
        # Scroll para o topo
        try:
            self.entry_nome.focus()
        except:
            pass
    
    def salvar_empresa_action(self):
        """Salva ou atualiza empresa"""
        nome = self.entry_nome.get().strip()
        cnpj = self.entry_cnpj.get().strip()
        usa_cert = self.var_usa_certificado.get()
        
        # Validações
        if not nome:
            messagebox.showwarning("Atenção", "Nome da empresa é obrigatório")
            self.entry_nome.focus()
            return
        
        if not cnpj:
            messagebox.showwarning("Atenção", "CNPJ é obrigatório")
            self.entry_cnpj.focus()
            return
        
        if not self.validador.validar_cnpj(cnpj):
            messagebox.showwarning(
                "CNPJ Inválido",
                "O CNPJ deve conter 14 dígitos numéricos"
            )
            self.entry_cnpj.focus()
            return
        
        # Dados específicos do tipo de login
        senha_portal = ""
        caminho_pfx = ""
        senha_pfx = ""
        
        if usa_cert:
            caminho_pfx = self.entry_pfx.get().strip()
            senha_pfx = self.entry_senha_pfx.get().strip()
            
            if not caminho_pfx:
                messagebox.showwarning("Atenção", "Selecione o arquivo .PFX")
                return
            
            if not os.path.exists(caminho_pfx):
                messagebox.showerror("Erro", "Arquivo .PFX não encontrado")
                return
            
            if not senha_pfx:
                messagebox.showwarning("Atenção", "Digite a senha do certificado")
                return
        else:
            senha_portal = self.entry_senha_portal.get().strip()
            if not senha_portal:
                messagebox.showwarning("Atenção", "Digite a senha do portal")
                return
        
        # Criar objeto empresa
        nova_empresa = {
            'nome': nome,
            'cnpj': cnpj,
            'usa_certificado': usa_cert,
            'senha': senha_portal,
            'caminho_pfx': caminho_pfx,
            'senha_pfx': senha_pfx
        }
        
        # Salvar ou atualizar
        if self.cnpj_em_edicao:
            # Atualizar existente
            for i, emp in enumerate(self.empresas):
                if str(emp['cnpj']) == str(self.cnpj_em_edicao):
                    self.empresas[i] = nova_empresa
                    break
            messagebox.showinfo("Sucesso", "Empresa atualizada com sucesso!")
            logger.info(f"Empresa atualizada: {nome}")
        else:
            # Verificar duplicata
            for emp in self.empresas:
                if str(emp['cnpj']) == str(cnpj):
                    messagebox.showerror("Erro", "CNPJ já cadastrado")
                    return
            
            # Adicionar nova
            self.empresas.append(nova_empresa)
            messagebox.showinfo("Sucesso", "Empresa cadastrada com sucesso!")
            logger.info(f"Empresa cadastrada: {nome}")
        
        self.salvar_tudo()
        self.atualizar_lista_visual()
        self.limpar_campos()
    
    def remover_empresa(self):
        """Remove empresa selecionada"""
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showinfo("Atenção", "Selecione uma ou mais empresas na lista")
            return
        
        # Preparar lista de empresas a remover
        empresas_remover = []
        for selected_item in selected_items:
            item = self.tree.item(selected_item)
            nome = item['values'][0]
            cnpj_formatado = str(item['values'][1])
            # Remover formatação
            cnpj = cnpj_formatado.replace('.', '').replace('/', '').replace('-', '')
            empresas_remover.append((nome, cnpj))
        
        # Confirmação
        if len(empresas_remover) == 1:
            msg = f"Deseja realmente remover a empresa:\n\n{empresas_remover[0][0]}\nCNPJ: {empresas_remover[0][1]}"
        else:
            msg = f"Deseja realmente remover {len(empresas_remover)} empresas selecionadas?"
        
        if messagebox.askyesno("Confirmar Remoção", msg):
            # Remover empresas
            cnpjs_remover = [cnpj for _, cnpj in empresas_remover]
            self.empresas = [
                emp for emp in self.empresas 
                if str(emp['cnpj']) not in cnpjs_remover
            ]
            self.salvar_tudo()
            self.atualizar_lista_visual()
            
            if len(empresas_remover) == 1:
                logger.info(f"Empresa removida: {empresas_remover[0][0]}")
                messagebox.showinfo("Sucesso", "Empresa removida")
            else:
                logger.info(f"{len(empresas_remover)} empresas removidas")
                messagebox.showinfo("Sucesso", f"{len(empresas_remover)} empresas removidas")
    
    def atualizar_lista_visual(self):
        """Atualiza a TreeView com lista de empresas"""
        # Limpar
        for i in self.tree.get_children():
            self.tree.delete(i)
        
        # Preencher
        for emp in self.empresas:
            tipo_txt = "🔒 Certificado" if emp.get('usa_certificado') else "🔑 Senha"
            cnpj_formatado = self.validador.formatar_cnpj(emp['cnpj'])
            self.tree.insert(
                '',
                'end',
                values=(emp.get('nome'), cnpj_formatado, tipo_txt)
            )
    
    def selecionar_todas_empresas(self):
        """Seleciona todas as empresas da lista"""
        # Obter todos os itens
        todos_itens = self.tree.get_children()
        
        if not todos_itens:
            messagebox.showinfo("Atenção", "Nenhuma empresa cadastrada")
            return
        
        # Selecionar todos
        self.tree.selection_set(todos_itens)
        
        # Atualizar label
        self.atualizar_contador_selecao()
        
        logger.info(f"Todas as {len(todos_itens)} empresas selecionadas")
    
    def desmarcar_todas_empresas(self):
        """Desmarca todas as empresas da lista"""
        self.tree.selection_remove(self.tree.selection())
        self.atualizar_contador_selecao()
        logger.info("Todas as seleções removidas")
    
    def atualizar_contador_selecao(self, event=None):
        """Atualiza o contador de empresas selecionadas"""
        qtd = len(self.tree.selection())
        
        if qtd == 0:
            self.lbl_selecionadas.config(
                text="💡 Use CTRL+Clique para múltiplas | SHIFT+Clique para faixa",
                fg="gray"
            )
        elif qtd == 1:
            self.lbl_selecionadas.config(
                text="✅ 1 empresa selecionada",
                fg="green"
            )
        else:
            self.lbl_selecionadas.config(
                text=f"✅ {qtd} empresas selecionadas",
                fg="blue"
            )
    
    # ========================================================================
    # EXECUÇÃO DO DOWNLOAD
    # ========================================================================
    
    def iniciar_thread(self):
        """Inicia thread de execução do download (suporta múltiplas empresas)"""
        # Validações
        path = self.path_download.get()
        if not path:
            messagebox.showerror("Erro", "Selecione a pasta raiz de downloads")
            return
        
        selected_items = self.tree.selection()
        if not selected_items:
            messagebox.showerror("Erro", "Selecione pelo menos uma empresa na lista")
            return
        
        # Validar que pelo menos um tipo foi selecionado
        baixar_emitidas = self.var_baixar_emitidas.get()
        baixar_recebidas = self.var_baixar_recebidas.get()
        
        if not baixar_emitidas and not baixar_recebidas:
            messagebox.showwarning(
                "Atenção",
                "Selecione pelo menos um tipo de NFS-e para baixar:\n"
                "• NFS-e EMITIDAS ou\n"
                "• NFS-e RECEBIDAS"
            )
            return
        
        # Validar competência se fornecida
        comp = self.competencia_filtro.get().strip()
        if comp:
            mes, ano = self.validador.validar_competencia(comp)
            if mes is None:
                messagebox.showerror(
                    "Erro",
                    "Competência inválida.\nFormato esperado: MM/AAAA"
                )
                return
        
        # Obter dados de TODAS as empresas selecionadas
        empresas_selecionadas = []
        for selected_item in selected_items:
            item = self.tree.item(selected_item)
            cnpj_sel = str(item['values'][1]).replace('.', '').replace('/', '').replace('-', '')
            
            emp_dados = next(
                (e for e in self.empresas if str(e.get('cnpj')) == cnpj_sel),
                None
            )
            
            if emp_dados:
                empresas_selecionadas.append(emp_dados)
        
        if not empresas_selecionadas:
            messagebox.showerror("Erro", "Nenhuma empresa válida selecionada")
            return
        
        # Confirmar se múltiplas empresas
        if len(empresas_selecionadas) > 1:
            resposta = messagebox.askyesno(
                "Múltiplas Empresas",
                f"Você selecionou {len(empresas_selecionadas)} empresas.\n\n"
                f"O download será feito em sequência.\n\n"
                f"Deseja continuar?"
            )
            if not resposta:
                return
        
        # Configurações
        tipo = self.tipo_download.get()
        usar_cache = self.var_usar_cache.get()
        modo_headless = self.var_modo_headless.get()
        
        # Resetar flag de cancelamento
        self.cancelar_flag = False
        
        # Desabilitar botão e habilitar cancelar
        self.btn_run.config(state="disabled")
        self.btn_cancelar_exec.config(state="normal")
        
        # Limpar log
        self.text_log.config(state='normal')
        self.text_log.delete(1.0, 'end')
        self.text_log.config(state='disabled')
        
        # Iniciar thread com LISTA de empresas
        self.thread_ativa = threading.Thread(
            target=self.executar_multiplas_empresas,
            args=(empresas_selecionadas, path, comp, tipo, usar_cache, modo_headless, baixar_emitidas, baixar_recebidas),
            daemon=True
        )
        self.thread_ativa.start()
        
        logger.info(f"Thread de download iniciada para {len(empresas_selecionadas)} empresa(s)")
    
    def cancelar_execucao(self):
        """Cancela execução em andamento"""
        if messagebox.askyesno(
            "Cancelar Execução",
            "Deseja realmente cancelar a execução em andamento?"
        ):
            self.cancelar_flag = True
            self.adicionar_log("⛔ Cancelamento solicitado...", "WARNING")
            logger.warning("Cancelamento solicitado pelo usuário")
    
    def executar_multiplas_empresas(
        self,
        lista_empresas,
        strCaminhoRaiz,
        strCompetenciaFiltro,
        strTipoDownload,
        usar_cache,
        modo_headless,
        baixar_emitidas,
        baixar_recebidas
    ):
        """
        Executa download para múltiplas empresas em sequência
        """
        total_empresas = len(lista_empresas)
        total_notas_geral = 0
        
        self.adicionar_log("="*60, "INFO")
        self.adicionar_log(f"🏢 PROCESSANDO {total_empresas} EMPRESA(S)", "INFO")
        self.adicionar_log("="*60, "INFO")
        
        for idx, empresa in enumerate(lista_empresas, 1):
            if self.cancelar_flag:
                self.adicionar_log("⛔ Processamento cancelado pelo usuário", "WARNING")
                break
            
            self.adicionar_log("\n" + "="*60, "INFO")
            self.adicionar_log(f"📊 EMPRESA {idx}/{total_empresas}", "INFO")
            self.adicionar_log("="*60 + "\n", "INFO")
            
            # Executar download para esta empresa
            try:
                count = self.executar_script_playwright_interno(
                    empresa,
                    strCaminhoRaiz,
                    strCompetenciaFiltro,
                    strTipoDownload,
                    usar_cache,
                    modo_headless,
                    baixar_emitidas,
                    baixar_recebidas
                )
                total_notas_geral += count
                
            except Exception as e:
                self.adicionar_log(f"❌ Erro ao processar {empresa.get('nome')}: {e}", "ERROR")
                logger.error(f"Erro empresa {empresa.get('nome')}: {e}\n{traceback.format_exc()}")
                continue
        
        # Mensagem final geral
        if total_empresas > 1:
            self.adicionar_log("\n" + "="*60, "SUCCESS")
            self.adicionar_log(
                f"✅ CONCLUÍDO! {total_empresas} empresas processadas\n"
                f"📊 Total geral: {total_notas_geral} notas baixadas",
                "SUCCESS"
            )
            self.adicionar_log("="*60, "SUCCESS")
            
            messagebox.showinfo(
                "Concluído",
                f"Processamento finalizado!\n\n"
                f"Empresas: {total_empresas}\n"
                f"Total de notas: {total_notas_geral}"
            )
        
        self.reset_ui()
    
    def executar_script_playwright_interno(
        self,
        dados_empresa: dict,
        strCaminhoRaiz: str,
        strCompetenciaFiltro: str,
        strTipoDownload: str,
        usar_cache: bool,
        modo_headless: bool,
        baixar_emitidas: bool,
        baixar_recebidas: bool
    ):
        """
        Execução principal do script de download com Playwright para UMA empresa
        Retorna quantidade de notas baixadas
        """
        try:
            # Extrair dados
            strNomeEmpresa = dados_empresa.get('nome')
            strCNPJ = dados_empresa.get('cnpj')
            usa_certificado = dados_empresa.get('usa_certificado', False)
            strSenhaPortal = dados_empresa.get('senha', '')
            caminho_pfx = dados_empresa.get('caminho_pfx', '')
            senha_pfx = dados_empresa.get('senha_pfx', '')
            
            self.adicionar_log("="*60, "INFO")
            self.adicionar_log(f"🏢 EMPRESA: {strNomeEmpresa}", "INFO")
            self.adicionar_log(f"📋 CNPJ: {self.validador.formatar_cnpj(strCNPJ)}", "INFO")
            self.adicionar_log(f"🔐 Tipo Acesso: {'Certificado' if usa_certificado else 'Senha'}", "INFO")
            self.adicionar_log(f"📅 Competência: {strCompetenciaFiltro or 'Últimos 30 dias'}", "INFO")
            self.adicionar_log(f"📥 Download: {strTipoDownload.upper()}", "INFO")
            self.adicionar_log("="*60, "INFO")
            
            self.atualizar_status("🔄 Preparando execução...", "blue")
            
            # Gerar períodos
            lista_periodos = GerenciadorPeriodos.gerar_periodos(strCompetenciaFiltro)
            self.adicionar_log(f"📆 Gerados {len(lista_periodos)} períodos de consulta", "INFO")
            
            # Preparar pasta
            nome_pasta_limpo = self.validador.limpar_nome_arquivo(strNomeEmpresa)
            caminho_empresa = os.path.join(strCaminhoRaiz, nome_pasta_limpo)
            Path(caminho_empresa).mkdir(parents=True, exist_ok=True)
            
            # Contadores
            contagem_clientes = {}
            count_total_baixados = 0
            count_ignorados_cache = 0
            
            # URL do portal
            strUrl = 'https://www.nfse.gov.br/EmissorNacional/Login?ReturnUrl=%2fEmissorNacional'
            
            # Iniciar Playwright
            with sync_playwright() as pw:
                self.adicionar_log("🌐 Iniciando navegador...", "INFO")
                self.atualizar_status("🌐 Iniciando navegador...", "blue")
                
                navegador = pw.chromium.launch(headless=modo_headless)
                contexto_args = {"ignore_https_errors": True}
                
                # Configurar certificado se necessário
                if usa_certificado:
                    if os.path.exists(caminho_pfx):
                        self.adicionar_log(f"🔒 Certificado: {os.path.basename(caminho_pfx)}", "INFO")
                        contexto_args["client_certificates"] = [{
                            "origin": "https://www.nfse.gov.br",
                            "pfxPath": caminho_pfx,
                            "passphrase": senha_pfx
                        }]
                    else:
                        raise Exception(f"Arquivo PFX não encontrado: {caminho_pfx}")
                
                contexto = navegador.new_context(**contexto_args)
                pagina = contexto.new_page()
                
                # Configurar timeout padrão
                pagina.set_default_timeout(30000)  # 30 segundos
                
                # LOGIN
                self.adicionar_log("🔑 Realizando login...", "INFO")
                self.atualizar_status("🔑 Fazendo login...", "blue")
                
                pagina.goto(strUrl, wait_until="networkidle")
                
                if usa_certificado:
                    # Login com certificado
                    try:
                        pagina.locator("a.img-certificado").click(timeout=5000)
                    except:
                        try:
                            pagina.get_by_text("Acesso via certificado digital").click(timeout=5000)
                        except:
                            self.adicionar_log("⚠️ Botão de certificado não encontrado, continuando...", "WARNING")
                    
                    # Aguardar login (verificação inteligente)
                    logado = False
                    for tentativa in range(30):  # 30 segundos
                        if self.cancelar_flag:
                            raise Exception("Operação cancelada pelo usuário")
                        
                        try:
                            if pagina.locator("text=Sair com segurança").is_visible(timeout=1000) or \
                               pagina.locator("text=Meus dados").is_visible(timeout=1000):
                                logado = True
                                self.adicionar_log("✅ Login detectado com sucesso!", "SUCCESS")
                                break
                        except:
                            pass
                    
                    if not logado:
                        self.adicionar_log("⚠️ Login não detectado claramente, tentando continuar...", "WARNING")
                
                else:
                    # Login com senha
                    try:
                        pagina.get_by_role("textbox", name="CPF/CNPJ").fill(strCNPJ)
                        pagina.get_by_role("textbox", name="Senha").fill(strSenhaPortal)
                        pagina.get_by_role("button", name="Entrar").click()
                        pagina.wait_for_load_state("networkidle", timeout=15000)
                        self.adicionar_log("✅ Login realizado", "SUCCESS")
                    except Exception as e:
                        raise Exception(f"Erro no login: {e}")
                
                # ============================================================
                # PROCESSAR NFS-E EMITIDAS
                # ============================================================
                if baixar_emitidas:
                    self.adicionar_log("\n" + "="*60, "INFO")
                    self.adicionar_log("📤 INICIANDO DOWNLOAD DE NFS-E EMITIDAS", "INFO")
                    self.adicionar_log("="*60, "INFO")
                    
                    count_emitidas = self._processar_emitidas(
                        pagina,
                        lista_periodos,
                        strCompetenciaFiltro,
                        strTipoDownload,
                        usar_cache,
                        caminho_empresa,
                        strNomeEmpresa,
                        contagem_clientes,
                        count_ignorados_cache
                    )
                    count_total_baixados += count_emitidas
                
                # ============================================================
                # PROCESSAR NFS-E RECEBIDAS
                # ============================================================
                if baixar_recebidas:
                    self.adicionar_log("\n" + "="*60, "INFO")
                    self.adicionar_log("📥 INICIANDO DOWNLOAD DE NFS-E RECEBIDAS", "INFO")
                    self.adicionar_log("="*60, "INFO")
                    
                    count_recebidas = self._processar_recebidas(
                        pagina,
                        lista_periodos,
                        strCompetenciaFiltro,
                        strTipoDownload,
                        usar_cache,
                        caminho_empresa,
                        strNomeEmpresa,
                        contagem_clientes,
                        count_ignorados_cache
                    )
                    count_total_baixados += count_recebidas
                
                # Fechar navegador
                navegador.close()
                
                # ============================================================
                # GERAR RELATÓRIOS EXCEL AUTOMÁTICOS
                # ============================================================
                if count_total_baixados > 0 and strCompetenciaFiltro:
                    self.adicionar_log("\n" + "="*60, "INFO")
                    self.adicionar_log("📊 GERANDO RELATÓRIOS EXCEL...", "INFO")
                    self.adicionar_log("="*60, "INFO")
                    
                    # Formatar competência para nome de pasta (MM-AAAA)
                    comp_formatada = strCompetenciaFiltro.replace('/', '-')
                    
                    relatorios_gerados = []
                    
                    # Gerar relatório de EMITIDAS
                    if baixar_emitidas:
                        try:
                            pasta_emitidas = os.path.join(caminho_empresa, "EMITIDAS", comp_formatada)
                            if os.path.exists(pasta_emitidas):
                                self.adicionar_log("📤 Processando XMLs de EMITIDAS...", "INFO")
                                xmls_emitidas = []
                                for root, dirs, files in os.walk(pasta_emitidas):
                                    if 'XML' in root:
                                        for file in files:
                                            if file.lower().endswith('.xml'):
                                                xmls_emitidas.append(os.path.join(root, file))
                                
                                if xmls_emitidas:
                                    nome_relatorio = f"Relatório_Emitidas_{comp_formatada}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    caminho_relatorio = os.path.join(caminho_empresa, "EMITIDAS", nome_relatorio)
                                    
                                    sucesso = GeradorRelatorioExcel.gerar_relatorio_emitidas(
                                        xmls_emitidas,
                                        caminho_relatorio
                                    )
                                    
                                    if sucesso:
                                        self.adicionar_log(f"  ✅ Relatório EMITIDAS gerado: {nome_relatorio}", "SUCCESS")
                                        relatorios_gerados.append(caminho_relatorio)
                                    else:
                                        self.adicionar_log(f"  ⚠️ Erro ao gerar relatório EMITIDAS", "WARNING")
                                else:
                                    self.adicionar_log("  ℹ️ Nenhum XML de EMITIDAS encontrado", "INFO")
                        except Exception as e_rel:
                            self.adicionar_log(f"  ❌ Erro ao gerar relatório EMITIDAS: {e_rel}", "ERROR")
                            logger.error(f"Erro relatório emitidas: {e_rel}")
                    
                    # Gerar relatório de RECEBIDAS
                    if baixar_recebidas:
                        try:
                            pasta_recebidas = os.path.join(caminho_empresa, "RECEBIDAS", comp_formatada)
                            if os.path.exists(pasta_recebidas):
                                self.adicionar_log("📥 Processando XMLs de RECEBIDAS...", "INFO")
                                xmls_recebidas = []
                                for root, dirs, files in os.walk(pasta_recebidas):
                                    if 'XML' in root:
                                        for file in files:
                                            if file.lower().endswith('.xml'):
                                                xmls_recebidas.append(os.path.join(root, file))
                                
                                if xmls_recebidas:
                                    nome_relatorio = f"Relatório_Recebidas_{comp_formatada}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                                    caminho_relatorio = os.path.join(caminho_empresa, "RECEBIDAS", nome_relatorio)
                                    
                                    sucesso = GeradorRelatorioExcel.gerar_relatorio_recebidas(
                                        xmls_recebidas,
                                        caminho_relatorio
                                    )
                                    
                                    if sucesso:
                                        self.adicionar_log(f"  ✅ Relatório RECEBIDAS gerado: {nome_relatorio}", "SUCCESS")
                                        relatorios_gerados.append(caminho_relatorio)
                                    else:
                                        self.adicionar_log(f"  ⚠️ Erro ao gerar relatório RECEBIDAS", "WARNING")
                                else:
                                    self.adicionar_log("  ℹ️ Nenhum XML de RECEBIDAS encontrado", "INFO")
                        except Exception as e_rel:
                            self.adicionar_log(f"  ❌ Erro ao gerar relatório RECEBIDAS: {e_rel}", "ERROR")
                            logger.error(f"Erro relatório recebidas: {e_rel}")
                    
                    if relatorios_gerados:
                        self.adicionar_log(f"\n📊 {len(relatorios_gerados)} relatório(s) Excel gerado(s) com sucesso!", "SUCCESS")
                
                # Mensagem final
                msg_emitidas = f"{count_total_baixados} nota(s)" if baixar_emitidas and baixar_recebidas else f"{count_total_baixados} nota(s)"
                
                msg_final = f"""
✅ DOWNLOAD CONCLUÍDO!

📊 Estatísticas:
  • Total baixado: {count_total_baixados} notas
  • Ignoradas (cache): {count_ignorados_cache} notas
  • Pasta: {caminho_empresa}
                """
                
                self.adicionar_log("\n" + "="*60, "SUCCESS")
                self.adicionar_log(msg_final, "SUCCESS")
                self.adicionar_log("="*60, "SUCCESS")
                
                self.atualizar_status("✅ Concluído com sucesso!", "green")
                
                logger.info(f"Download concluído: {count_total_baixados} notas")
                
                # Retornar contagem
                return count_total_baixados
        
        except Exception as e:
            msg_erro = f"Erro durante execução:\n{str(e)}"
            self.adicionar_log(f"\n❌ ERRO CRÍTICO: {e}", "ERROR")
            self.atualizar_status("❌ Erro na execução", "red")
            logger.error(f"Erro crítico:\n{traceback.format_exc()}")
            return 0
    

    def _processar_emitidas(
        self,
        pagina,
        lista_periodos,
        strCompetenciaFiltro,
        strTipoDownload,
        usar_cache,
        caminho_empresa,
        strNomeEmpresa,
        contagem_clientes,
        count_ignorados_cache_inicial
    ):
        """Processa download de NFS-e EMITIDAS"""
        count_total_baixados = 0
        count_ignorados_cache = count_ignorados_cache_inicial
        
        try:
            # Navegar para Notas Emitidas
            self.adicionar_log("📋 Navegando para Notas Emitidas...", "INFO")
            pagina.goto(
                "https://www.nfse.gov.br/EmissorNacional/Notas/Emitidas",
                wait_until="networkidle"
            )
            time.sleep(2)
            
            # Criar pasta EMITIDAS
            caminho_emitidas = os.path.join(caminho_empresa, "EMITIDAS")
            Path(caminho_emitidas).mkdir(parents=True, exist_ok=True)
            
            # LOOP DE PERÍODOS
            for idx_periodo, (p_inicio, p_fim) in enumerate(lista_periodos):
                if self.cancelar_flag:
                    raise Exception("Operação cancelada pelo usuário")
                
                periodo_str = f"{p_inicio} a {p_fim}" if p_inicio else "Últimos 30 dias"
                self.adicionar_log(f"\n📅 Período {idx_periodo+1}/{len(lista_periodos)}: {periodo_str}", "INFO")
                self.atualizar_status(f"🔍 Consultando EMITIDAS: {periodo_str}", "blue")
                
                # Aplicar filtro de data
                if p_inicio and p_fim:
                    try:
                        # Seletor data inicial
                        sel_inicio = "input#datainicio"
                        try:
                            pagina.wait_for_selector(sel_inicio, timeout=5000)
                        except:
                            sel_inicio = "input[name='datainicio']"
                            pagina.wait_for_selector(sel_inicio, timeout=5000)
                        
                        # Preencher data inicial
                        pagina.locator(sel_inicio).click()
                        pagina.locator(sel_inicio).press("Control+a")
                        pagina.locator(sel_inicio).fill(p_inicio)
                        
                        # Seletor data final
                        sel_fim = "input#datafim"
                        try:
                            pagina.wait_for_selector(sel_fim, timeout=5000)
                        except:
                            sel_fim = "input[name='datafim']"
                            pagina.wait_for_selector(sel_fim, timeout=5000)
                        
                        # Preencher data final
                        pagina.locator(sel_fim).click()
                        pagina.locator(sel_fim).press("Control+a")
                        pagina.locator(sel_fim).fill(p_fim)
                        
                        # Clicar em filtrar
                        pagina.locator("button").filter(has_text="Filtrar").first.click()
                        pagina.wait_for_load_state("networkidle", timeout=15000)
                        time.sleep(2)
                        
                        self.adicionar_log("✅ Filtro aplicado", "SUCCESS")
                    except Exception as e:
                        self.adicionar_log(f"❌ Erro ao aplicar filtro: {e}", "ERROR")
                        logger.error(f"Erro filtro: {e}\n{traceback.format_exc()}")
                        continue
                
                # Verificar total de notas
                try:
                    # OTIMIZAÇÃO: Verificar primeiro se tem a mensagem "sem registros"
                    # Isso evita timeout de 10 segundos quando não há notas
                    try:
                        sem_registros = pagina.locator(".sem-registros")
                        if sem_registros.is_visible(timeout=1000):  # Apenas 1 segundo
                            self.adicionar_log("⚡ Nenhuma nota EMITIDA (detecção rápida)", "INFO")
                            continue
                    except:
                        pass  # Continua tentando pelo método normal
                    
                    # Método normal: verificar "Total de X registros"
                    texto_total = pagina.get_by_text('Total de ').inner_text()
                    total_notas_site = int(texto_total.split(' ')[2])
                except:
                    total_notas_site = 0
                
                if total_notas_site == 0:
                    self.adicionar_log("ℹ️ Nenhuma nota EMITIDA encontrada neste período", "INFO")
                    continue
                
                self.adicionar_log(f"📊 Total de notas EMITIDAS: {total_notas_site}", "INFO")
                
                # LOOP DE PAGINAÇÃO
                count_processados_periodo = 0
                
                while count_processados_periodo < total_notas_site:
                    if self.cancelar_flag:
                        raise Exception("Operação cancelada pelo usuário")
                    
                    # Aguardar linhas da tabela
                    try:
                        pagina.wait_for_selector("tbody tr", timeout=10000)
                    except:
                        self.adicionar_log("⚠️ Timeout aguardando linhas", "WARNING")
                        break
                    
                    row_locator = pagina.locator("tbody tr")
                    count_na_pagina = row_locator.count()
                    
                    if count_na_pagina == 0:
                        self.adicionar_log("ℹ️ Nenhuma linha na página", "INFO")
                        break
                    
                    self.adicionar_log(
                        f"📄 Página atual: {count_na_pagina} notas | "
                        f"Processadas: {count_processados_periodo}/{total_notas_site}",
                        "INFO"
                    )
                    
                    # LOOP DE LINHAS
                    for i in range(count_na_pagina):
                        if self.cancelar_flag:
                            raise Exception("Operação cancelada pelo usuário")
                        
                        linha = row_locator.nth(i)
                        
                        try:
                            # Competência
                            txt_competencia = linha.locator(".td-competencia").inner_text().strip()
                            
                            # Filtrar por competência se especificado
                            if strCompetenciaFiltro and txt_competencia != strCompetenciaFiltro:
                                count_processados_periodo += 1
                                continue
                            
                            # Tomador (quem RECEBEU o serviço)
                            txt_tomador = linha.locator(".td-texto-grande").inner_text()
                            if ' - ' in txt_tomador:
                                nome_tomador = txt_tomador.split(' - ')[-1].strip()
                            elif '-' in txt_tomador:
                                nome_tomador = txt_tomador.split('-')[-1].strip()
                            else:
                                nome_tomador = txt_tomador.strip()
                            
                            nome_tomador = self.validador.limpar_nome_arquivo(nome_tomador)
                            
                            # Status
                            try:
                                strStatus = linha.locator(".td-situacao > img").get_attribute('data-original-title')
                                if not strStatus:
                                    strStatus = "Indefinido"
                                strStatus = strStatus.replace('/', '-').strip()
                            except:
                                strStatus = "StatusNaoLido"
                            
                            # Verificar cache
                            if usar_cache:
                                hash_nota = self.cache.gerar_hash(
                                    strNomeEmpresa,
                                    txt_competencia,
                                    nome_tomador
                                )
                                
                                if self.cache.ja_baixado(hash_nota):
                                    count_ignorados_cache += 1
                                    count_processados_periodo += 1
                                    continue
                            
                            # Preparar pastas (EMITIDAS/)
                            pasta_comp = txt_competencia.replace('/', '-')
                            dir_base = os.path.join(caminho_emitidas, pasta_comp, strStatus)
                            dir_xml = os.path.join(dir_base, "XML")
                            dir_pdf = os.path.join(dir_base, "PDF")
                            
                            self.adicionar_log(
                                f"⬇️ EMITIDA para: {nome_tomador[:30]}... | {strStatus}",
                                "INFO"
                            )
                            self.atualizar_status(
                                f"⬇️ Baixando EMITIDA: {nome_tomador[:40]}...",
                                "green"
                            )
                            
                            # Abrir menu de opções (três pontinhos)
                            try:
                                botao_opcoes = linha.locator(".icone-trigger").first
                                botao_opcoes.click(timeout=5000)
                                pagina.wait_for_timeout(500)
                            except Exception as e:
                                self.adicionar_log(f"⚠️ Erro ao abrir menu: {e}", "WARNING")
                                count_processados_periodo += 1
                                continue
                            
                            # DOWNLOAD XML
                            if strTipoDownload in ['xml', 'ambos']:
                                Path(dir_xml).mkdir(parents=True, exist_ok=True)
                                try:
                                    link_xml = linha.locator("a").filter(has_text="Download XML").first
                                    
                                    with pagina.expect_download(timeout=30000) as dl_info:
                                        link_xml.click()
                                    
                                    dl = dl_info.value
                                    caminho_xml = os.path.join(dir_xml, dl.suggested_filename)
                                    dl.save_as(caminho_xml)
                                    
                                    self.adicionar_log(f"  ✅ XML salvo", "SUCCESS")
                                except Exception as e:
                                    self.adicionar_log(f"  ❌ Erro XML: {e}", "ERROR")
                                    logger.error(f"Erro download XML: {e}")
                            
                            # DOWNLOAD PDF
                            if strTipoDownload in ['pdf', 'ambos']:
                                Path(dir_pdf).mkdir(parents=True, exist_ok=True)
                                try:
                                    # SEMPRE reabrir o menu para PDF (o XML fecha o menu)
                                    try:
                                        botao_opcoes = linha.locator(".icone-trigger").first
                                        botao_opcoes.click(timeout=5000)
                                        pagina.wait_for_timeout(800)  # Aguardar menu abrir completamente
                                    except Exception as e_menu:
                                        self.adicionar_log(f"  ⚠️ Erro ao reabrir menu para PDF: {e_menu}", "WARNING")
                                        raise
                                    
                                    link_pdf = linha.locator("a").filter(has_text="Download DANFS-e").first
                                    
                                    # Verificar se link está visível
                                    if not link_pdf.is_visible(timeout=2000):
                                        self.adicionar_log(f"  ⚠️ Link PDF não visível após abrir menu", "WARNING")
                                        raise Exception("Link PDF não visível")
                                    
                                    # Nome do PDF
                                    qtd = contagem_clientes.get(nome_tomador, 0) + 1
                                    contagem_clientes[nome_tomador] = qtd
                                    nome_pdf = f"{nome_tomador}_{qtd}.pdf"
                                    
                                    with pagina.expect_download(timeout=30000) as dl_info:
                                        link_pdf.click()
                                    
                                    dl = dl_info.value
                                    caminho_pdf = os.path.join(dir_pdf, nome_pdf)
                                    dl.save_as(caminho_pdf)
                                    
                                    self.adicionar_log(f"  ✅ PDF salvo: {nome_pdf}", "SUCCESS")
                                except Exception as e:
                                    self.adicionar_log(f"  ❌ Erro PDF: {e}", "ERROR")
                                    logger.error(f"Erro download PDF: {e}")
                            
                            # Registrar no cache
                            if usar_cache:
                                self.cache.registrar_download(hash_nota)
                            
                            count_total_baixados += 1
                            
                        except Exception as e_linha:
                            self.adicionar_log(f"❌ Erro ao processar linha: {e_linha}", "ERROR")
                            logger.error(f"Erro linha: {e_linha}\n{traceback.format_exc()}")
                        
                        count_processados_periodo += 1
                    
                    # Verificar se há próxima página
                    if count_processados_periodo < total_notas_site:
                        try:
                            btn_prox = None
                            metodo_usado = ""
                            
                            # Tentativa 1: Por ícone Font Awesome
                            try:
                                btn_prox = pagina.locator("a[data-original-title='Próxima'] i.fa-angle-right").locator("..").first
                                if btn_prox.is_visible(timeout=1000):
                                    metodo_usado = "fa-angle-right (tooltip Próxima)"
                                else:
                                    btn_prox = None
                            except:
                                pass
                            
                            # Tentativa 2: Genérico
                            if not btn_prox:
                                try:
                                    btn_prox = pagina.locator("li:not(.disabled) a i.fa-angle-right").locator("..").first
                                    if btn_prox.is_visible(timeout=1000):
                                        metodo_usado = "fa-angle-right (genérico)"
                                    else:
                                        btn_prox = None
                                except:
                                    pass
                            
                            if btn_prox and btn_prox.is_visible():
                                self.adicionar_log(f"➡️ Próxima página... (método: {metodo_usado})", "INFO")
                                btn_prox.click()
                                pagina.wait_for_load_state("networkidle", timeout=15000)
                                time.sleep(2)
                            else:
                                self.adicionar_log("ℹ️ Não há mais páginas", "INFO")
                                break
                        except Exception as e:
                            self.adicionar_log(f"ℹ️ Fim da paginação: {e}", "INFO")
                            break
                    else:
                        break
            
            self.adicionar_log(f"\n✅ EMITIDAS: {count_total_baixados} notas baixadas", "SUCCESS")
            return count_total_baixados
            
        except Exception as e:
            self.adicionar_log(f"❌ Erro ao processar EMITIDAS: {e}", "ERROR")
            logger.error(f"Erro EMITIDAS: {e}\n{traceback.format_exc()}")
            return count_total_baixados
    
    def _processar_recebidas(
        self,
        pagina,
        lista_periodos,
        strCompetenciaFiltro,
        strTipoDownload,
        usar_cache,
        caminho_empresa,
        strNomeEmpresa,
        contagem_clientes,
        count_ignorados_cache_inicial
    ):
        """Processa download de NFS-e RECEBIDAS"""
        count_total_baixados = 0
        count_ignorados_cache = count_ignorados_cache_inicial
        
        try:
            # Navegar para Notas Recebidas
            self.adicionar_log("📋 Navegando para Notas Recebidas...", "INFO")
            pagina.goto(
                "https://www.nfse.gov.br/EmissorNacional/Notas/Recebidas?executar=1",
                wait_until="networkidle"
            )
            time.sleep(2)
            
            # Criar pasta RECEBIDAS
            caminho_recebidas = os.path.join(caminho_empresa, "RECEBIDAS")
            Path(caminho_recebidas).mkdir(parents=True, exist_ok=True)
            
            # LOOP DE PERÍODOS
            for idx_periodo, (p_inicio, p_fim) in enumerate(lista_periodos):
                if self.cancelar_flag:
                    raise Exception("Operação cancelada pelo usuário")
                
                periodo_str = f"{p_inicio} a {p_fim}" if p_inicio else "Últimos 30 dias"
                self.adicionar_log(f"\n📅 Período {idx_periodo+1}/{len(lista_periodos)}: {periodo_str}", "INFO")
                self.atualizar_status(f"🔍 Consultando RECEBIDAS: {periodo_str}", "blue")
                
                # Aplicar filtro de data
                if p_inicio and p_fim:
                    try:
                        # Seletor data inicial
                        sel_inicio = "input#datainicio"
                        try:
                            pagina.wait_for_selector(sel_inicio, timeout=5000)
                        except:
                            sel_inicio = "input[name='datainicio']"
                            pagina.wait_for_selector(sel_inicio, timeout=5000)
                        
                        # Preencher data inicial
                        pagina.locator(sel_inicio).click()
                        pagina.locator(sel_inicio).press("Control+a")
                        pagina.locator(sel_inicio).fill(p_inicio)
                        
                        # Seletor data final
                        sel_fim = "input#datafim"
                        try:
                            pagina.wait_for_selector(sel_fim, timeout=5000)
                        except:
                            sel_fim = "input[name='datafim']"
                            pagina.wait_for_selector(sel_fim, timeout=5000)
                        
                        # Preencher data final
                        pagina.locator(sel_fim).click()
                        pagina.locator(sel_fim).press("Control+a")
                        pagina.locator(sel_fim).fill(p_fim)
                        
                        # Clicar em filtrar
                        pagina.locator("button").filter(has_text="Filtrar").first.click()
                        pagina.wait_for_load_state("networkidle", timeout=15000)
                        time.sleep(2)
                        
                        self.adicionar_log("✅ Filtro aplicado", "SUCCESS")
                    except Exception as e:
                        self.adicionar_log(f"❌ Erro ao aplicar filtro: {e}", "ERROR")
                        logger.error(f"Erro filtro: {e}\n{traceback.format_exc()}")
                        continue
                
                # Verificar total de notas
                try:
                    # OTIMIZAÇÃO: Verificar primeiro se tem a mensagem "sem registros"
                    # Isso evita timeout de 10 segundos quando não há notas
                    try:
                        sem_registros = pagina.locator(".sem-registros")
                        if sem_registros.is_visible(timeout=1000):  # Apenas 1 segundo
                            self.adicionar_log("⚡ Nenhuma nota RECEBIDA (detecção rápida)", "INFO")
                            continue
                    except:
                        pass  # Continua tentando pelo método normal
                    
                    # Método normal: verificar "Total de X registros"
                    texto_total = pagina.get_by_text('Total de ').inner_text()
                    total_notas_site = int(texto_total.split(' ')[2])
                except:
                    total_notas_site = 0
                
                if total_notas_site == 0:
                    self.adicionar_log("ℹ️ Nenhuma nota RECEBIDA encontrada neste período", "INFO")
                    continue
                
                self.adicionar_log(f"📊 Total de notas RECEBIDAS: {total_notas_site}", "INFO")
                
                # LOOP DE PAGINAÇÃO
                count_processados_periodo = 0
                
                while count_processados_periodo < total_notas_site:
                    if self.cancelar_flag:
                        raise Exception("Operação cancelada pelo usuário")
                    
                    # Aguardar linhas da tabela
                    try:
                        pagina.wait_for_selector("tbody tr", timeout=10000)
                    except:
                        self.adicionar_log("⚠️ Timeout aguardando linhas", "WARNING")
                        break
                    
                    row_locator = pagina.locator("tbody tr")
                    count_na_pagina = row_locator.count()
                    
                    if count_na_pagina == 0:
                        self.adicionar_log("ℹ️ Nenhuma linha na página", "INFO")
                        break
                    
                    self.adicionar_log(
                        f"📄 Página atual: {count_na_pagina} notas | "
                        f"Processadas: {count_processados_periodo}/{total_notas_site}",
                        "INFO"
                    )
                    
                    # LOOP DE LINHAS
                    for i in range(count_na_pagina):
                        if self.cancelar_flag:
                            raise Exception("Operação cancelada pelo usuário")
                        
                        linha = row_locator.nth(i)
                        
                        try:
                            # Competência
                            txt_competencia = linha.locator(".td-competencia").inner_text().strip()
                            
                            # Filtrar por competência se especificado
                            if strCompetenciaFiltro and txt_competencia != strCompetenciaFiltro:
                                count_processados_periodo += 1
                                continue
                            
                            # Prestador (quem EMITIU a nota) - DIFERENÇA PRINCIPAL
                            txt_prestador = linha.locator(".td-texto-grande").inner_text()
                            if ' - ' in txt_prestador:
                                nome_prestador = txt_prestador.split(' - ')[-1].strip()
                            elif '-' in txt_prestador:
                                nome_prestador = txt_prestador.split('-')[-1].strip()
                            else:
                                nome_prestador = txt_prestador.strip()
                            
                            nome_prestador = self.validador.limpar_nome_arquivo(nome_prestador)
                            
                            # Status
                            try:
                                strStatus = linha.locator(".td-situacao > img").get_attribute('data-original-title')
                                if not strStatus:
                                    strStatus = "Indefinido"
                                strStatus = strStatus.replace('/', '-').strip()
                            except:
                                strStatus = "StatusNaoLido"
                            
                            # Verificar cache (usando PRESTADOR ao invés de tomador)
                            if usar_cache:
                                hash_nota = self.cache.gerar_hash(
                                    strNomeEmpresa,
                                    txt_competencia,
                                    nome_prestador
                                )
                                
                                if self.cache.ja_baixado(hash_nota):
                                    count_ignorados_cache += 1
                                    count_processados_periodo += 1
                                    continue
                            
                            # Preparar pastas (RECEBIDAS/Competencia/PDF e XML diretamente)
                            pasta_comp = txt_competencia.replace('/', '-')
                            dir_base = os.path.join(caminho_recebidas, pasta_comp)
                            dir_xml = os.path.join(dir_base, "XML")
                            dir_pdf = os.path.join(dir_base, "PDF")
                            
                            self.adicionar_log(
                                f"⬇️ RECEBIDA de: {nome_prestador[:30]}... | {strStatus}",
                                "INFO"
                            )
                            self.atualizar_status(
                                f"⬇️ Baixando RECEBIDA: {nome_prestador[:40]}...",
                                "green"
                            )
                            
                            # Abrir menu de opções (três pontinhos)
                            try:
                                botao_opcoes = linha.locator(".icone-trigger").first
                                botao_opcoes.click(timeout=5000)
                                pagina.wait_for_timeout(500)
                            except Exception as e:
                                self.adicionar_log(f"⚠️ Erro ao abrir menu: {e}", "WARNING")
                                count_processados_periodo += 1
                                continue
                            
                            # DOWNLOAD XML
                            if strTipoDownload in ['xml', 'ambos']:
                                Path(dir_xml).mkdir(parents=True, exist_ok=True)
                                try:
                                    link_xml = linha.locator("a").filter(has_text="Download XML").first
                                    
                                    with pagina.expect_download(timeout=30000) as dl_info:
                                        link_xml.click()
                                    
                                    dl = dl_info.value
                                    caminho_xml = os.path.join(dir_xml, dl.suggested_filename)
                                    dl.save_as(caminho_xml)
                                    
                                    self.adicionar_log(f"  ✅ XML salvo", "SUCCESS")
                                except Exception as e:
                                    self.adicionar_log(f"  ❌ Erro XML: {e}", "ERROR")
                                    logger.error(f"Erro download XML: {e}")
                            
                            # DOWNLOAD PDF
                            if strTipoDownload in ['pdf', 'ambos']:
                                Path(dir_pdf).mkdir(parents=True, exist_ok=True)
                                try:
                                    # SEMPRE reabrir o menu para PDF (o XML fecha o menu)
                                    try:
                                        botao_opcoes = linha.locator(".icone-trigger").first
                                        botao_opcoes.click(timeout=5000)
                                        pagina.wait_for_timeout(800)  # Aguardar menu abrir completamente
                                    except Exception as e_menu:
                                        self.adicionar_log(f"  ⚠️ Erro ao reabrir menu para PDF: {e_menu}", "WARNING")
                                        raise
                                    
                                    link_pdf = linha.locator("a").filter(has_text="Download DANFS-e").first
                                    
                                    # Verificar se link está visível
                                    if not link_pdf.is_visible(timeout=2000):
                                        self.adicionar_log(f"  ⚠️ Link PDF não visível após abrir menu", "WARNING")
                                        raise Exception("Link PDF não visível")
                                    
                                    # Nome do PDF
                                    qtd = contagem_clientes.get(nome_prestador, 0) + 1
                                    contagem_clientes[nome_prestador] = qtd
                                    nome_pdf = f"{nome_prestador}_{qtd}.pdf"
                                    
                                    with pagina.expect_download(timeout=30000) as dl_info:
                                        link_pdf.click()
                                    
                                    dl = dl_info.value
                                    caminho_pdf = os.path.join(dir_pdf, nome_pdf)
                                    dl.save_as(caminho_pdf)
                                    
                                    self.adicionar_log(f"  ✅ PDF salvo: {nome_pdf}", "SUCCESS")
                                except Exception as e:
                                    self.adicionar_log(f"  ❌ Erro PDF: {e}", "ERROR")
                                    logger.error(f"Erro download PDF: {e}")
                            
                            # Registrar no cache
                            if usar_cache:
                                self.cache.registrar_download(hash_nota)
                            
                            count_total_baixados += 1
                            
                        except Exception as e_linha:
                            self.adicionar_log(f"❌ Erro ao processar linha: {e_linha}", "ERROR")
                            logger.error(f"Erro linha: {e_linha}\n{traceback.format_exc()}")
                        
                        count_processados_periodo += 1
                    
                    # Verificar se há próxima página (RECEBIDAS não tem paginação no HTML que vi)
                    # Mas vou deixar a lógica caso tenha
                    if count_processados_periodo < total_notas_site:
                        try:
                            btn_prox = None
                            metodo_usado = ""
                            
                            # Tentativa 1: Por ícone Font Awesome
                            try:
                                btn_prox = pagina.locator("a[data-original-title='Próxima'] i.fa-angle-right").locator("..").first
                                if btn_prox.is_visible(timeout=1000):
                                    metodo_usado = "fa-angle-right (tooltip Próxima)"
                                else:
                                    btn_prox = None
                            except:
                                pass
                            
                            # Tentativa 2: Genérico
                            if not btn_prox:
                                try:
                                    btn_prox = pagina.locator("li:not(.disabled) a i.fa-angle-right").locator("..").first
                                    if btn_prox.is_visible(timeout=1000):
                                        metodo_usado = "fa-angle-right (genérico)"
                                    else:
                                        btn_prox = None
                                except:
                                    pass
                            
                            if btn_prox and btn_prox.is_visible():
                                self.adicionar_log(f"➡️ Próxima página... (método: {metodo_usado})", "INFO")
                                btn_prox.click()
                                pagina.wait_for_load_state("networkidle", timeout=15000)
                                time.sleep(2)
                            else:
                                self.adicionar_log("ℹ️ Não há mais páginas", "INFO")
                                break
                        except Exception as e:
                            self.adicionar_log(f"ℹ️ Fim da paginação: {e}", "INFO")
                            break
                    else:
                        break
            
            self.adicionar_log(f"\n✅ RECEBIDAS: {count_total_baixados} notas baixadas", "SUCCESS")
            return count_total_baixados
            
        except Exception as e:
            self.adicionar_log(f"❌ Erro ao processar RECEBIDAS: {e}", "ERROR")
            logger.error(f"Erro RECEBIDAS: {e}\n{traceback.format_exc()}")
            return count_total_baixados

    
    def gerar_relatorios_manual(self):
        """Gera relatórios Excel manualmente a partir de XMLs selecionados"""
        try:
            # Criar janela de diálogo personalizada
            dialog = tk.Toplevel(self.root)
            dialog.title("Gerar Relatórios Excel")
            dialog.geometry("700x600")  # Aumentado de 650x450
            dialog.transient(self.root)
            dialog.grab_set()
            
            # Centralizar janela
            dialog.update_idletasks()
            x = (dialog.winfo_screenwidth() // 2) - (700 // 2)
            y = (dialog.winfo_screenheight() // 2) - (600 // 2)
            dialog.geometry(f"700x600+{x}+{y}")
            
            # Título
            tk.Label(
                dialog,
                text="📊 Gerar Relatórios Excel",
                font=("Arial", 14, "bold")
            ).pack(pady=10)
            
            # Instruções (MAIS COMPACTO)
            frame_instrucoes = tk.LabelFrame(dialog, text="Como Funciona", padx=10, pady=5)
            frame_instrucoes.pack(fill="x", padx=20, pady=5)
            
            tk.Label(
                frame_instrucoes,
                text="1. Clique em 'Selecionar XMLs' e escolha os arquivos\n"
                     "2. Escolha o tipo: EMITIDAS ou RECEBIDAS\n"
                     "3. Clique em 'GERAR RELATÓRIO' e escolha onde salvar",
                justify="left",
                font=("Arial", 9)
            ).pack(anchor="w")
            
            # Variáveis
            xmls_selecionados = []
            
            # Frame seleção de XMLs (ALTURA FIXA)
            frame_xmls = tk.LabelFrame(dialog, text="XMLs Selecionados", padx=10, pady=10)
            frame_xmls.pack(fill="both", expand=True, padx=20, pady=10)
            
            # Label contador
            lbl_contador = tk.Label(
                frame_xmls,
                text="Nenhum XML selecionado",
                font=("Arial", 10, "bold"),
                fg="gray"
            )
            lbl_contador.pack(pady=5)
            
            # Listbox para mostrar XMLs (ALTURA FIXA)
            frame_lista = tk.Frame(frame_xmls)
            frame_lista.pack(fill="both", expand=True, pady=5)
            
            scrollbar = tk.Scrollbar(frame_lista)
            scrollbar.pack(side="right", fill="y")
            
            listbox = tk.Listbox(
                frame_lista,
                yscrollcommand=scrollbar.set,
                font=("Consolas", 8),
                height=10  # Altura fixa
            )
            listbox.pack(side="left", fill="both", expand=True)
            scrollbar.config(command=listbox.yview)
            
            def selecionar_xmls():
                arquivos = filedialog.askopenfilenames(
                    title="Selecione os arquivos XML",
                    filetypes=[
                        ("Arquivos XML", "*.xml"),
                        ("Todos os arquivos", "*.*")
                    ],
                    initialdir=self.path_download.get() if self.path_download.get() else os.path.expanduser("~")
                )
                
                if arquivos:
                    xmls_selecionados.clear()
                    xmls_selecionados.extend(arquivos)
                    
                    # Atualizar listbox
                    listbox.delete(0, tk.END)
                    for xml in xmls_selecionados:
                        nome_arquivo = os.path.basename(xml)
                        listbox.insert(tk.END, nome_arquivo)
                    
                    # Atualizar contador
                    qtd = len(xmls_selecionados)
                    lbl_contador.config(
                        text=f"✅ {qtd} arquivo(s) XML selecionado(s)",
                        fg="green"
                    )
            
            def limpar_selecao():
                xmls_selecionados.clear()
                listbox.delete(0, tk.END)
                lbl_contador.config(
                    text="Nenhum XML selecionado",
                    fg="gray"
                )
            
            # Botões de seleção
            frame_botoes_xml = tk.Frame(frame_xmls)
            frame_botoes_xml.pack(fill="x", pady=5)
            
            tk.Button(
                frame_botoes_xml,
                text="📁 Selecionar XMLs...",
                command=selecionar_xmls,
                cursor="hand2",
                font=("Arial", 9),
                bg="#E3F2FD"
            ).pack(side="left", padx=(0, 5))
            
            tk.Button(
                frame_botoes_xml,
                text="🗑️ Limpar",
                command=limpar_selecao,
                cursor="hand2",
                font=("Arial", 9)
            ).pack(side="left")
            
            # Frame tipo de relatório (COMPACTO)
            frame_tipo = tk.LabelFrame(dialog, text="Tipo de Relatório", padx=10, pady=5)
            frame_tipo.pack(fill="x", padx=20, pady=10)
            
            var_tipo = tk.StringVar(value="emitidas")
            
            tk.Radiobutton(
                frame_tipo,
                text="📤 NFS-e EMITIDAS (você emitiu)",
                variable=var_tipo,
                value="emitidas",
                font=("Arial", 9)
            ).pack(anchor="w", pady=2)
            
            tk.Radiobutton(
                frame_tipo,
                text="📥 NFS-e RECEBIDAS (você recebeu)",
                variable=var_tipo,
                value="recebidas",
                font=("Arial", 9)
            ).pack(anchor="w", pady=2)
            
            # Frame botões (SEMPRE VISÍVEL)
            frame_botoes = tk.Frame(dialog, bg="#f0f0f0", relief="raised", bd=1)
            frame_botoes.pack(fill="x", padx=0, pady=0, side="bottom")
            
            def processar():
                if not xmls_selecionados:
                    messagebox.showwarning("Atenção", "Selecione pelo menos um arquivo XML")
                    return
                
                tipo = var_tipo.get()
                
                # Perguntar onde salvar
                nome_padrao = f"Relatório_{'Emitidas' if tipo == 'emitidas' else 'Recebidas'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                arquivo_destino = filedialog.asksaveasfilename(
                    title="Salvar Relatório Como",
                    defaultextension=".xlsx",
                    filetypes=[("Arquivo Excel", "*.xlsx")],
                    initialfile=nome_padrao,
                    initialdir=self.path_download.get() if self.path_download.get() else os.path.expanduser("~")
                )
                
                if not arquivo_destino:
                    return  # Usuário cancelou
                
                # Fechar dialog
                dialog.destroy()
                
                # Processar em thread
                threading.Thread(
                    target=self._gerar_relatorio_xmls_selecionados,
                    args=(xmls_selecionados.copy(), tipo, arquivo_destino),
                    daemon=True
                ).start()
            
            # Espaçamento interno
            tk.Frame(frame_botoes, height=10).pack()
            
            # Botões em frame interno
            frame_botoes_interno = tk.Frame(frame_botoes, bg="#f0f0f0")
            frame_botoes_interno.pack(pady=10)
            
            tk.Button(
                frame_botoes_interno,
                text="✅ GERAR RELATÓRIO",
                font=("Arial", 11, "bold"),
                bg="#4CAF50",
                fg="white",
                command=processar,
                cursor="hand2",
                width=25,
                height=2
            ).pack(side="left", padx=10)
            
            tk.Button(
                frame_botoes_interno,
                text="❌ Cancelar",
                font=("Arial", 11),
                command=dialog.destroy,
                cursor="hand2",
                width=15,
                height=2
            ).pack(side="left", padx=10)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao abrir janela:\n{e}")
            logger.error(f"Erro gerar_relatorios_manual: {e}\n{traceback.format_exc()}")
    
    def _gerar_relatorio_xmls_selecionados(self, lista_xmls, tipo, arquivo_destino):
        """Gera relatório a partir de XMLs selecionados manualmente"""
        try:
            self.adicionar_log("\n" + "="*60, "INFO")
            self.adicionar_log("📊 GERANDO RELATÓRIO MANUAL", "INFO")
            self.adicionar_log(f"📁 XMLs selecionados: {len(lista_xmls)}", "INFO")
            self.adicionar_log(f"📋 Tipo: {tipo.upper()}", "INFO")
            self.adicionar_log("="*60, "INFO")
            
            # Gerar relatório
            sucesso = False
            if tipo == "emitidas":
                self.adicionar_log("📤 Processando XMLs como EMITIDAS...", "INFO")
                sucesso = GeradorRelatorioExcel.gerar_relatorio_emitidas(
                    lista_xmls,
                    arquivo_destino
                )
            else:
                self.adicionar_log("📥 Processando XMLs como RECEBIDAS...", "INFO")
                sucesso = GeradorRelatorioExcel.gerar_relatorio_recebidas(
                    lista_xmls,
                    arquivo_destino
                )
            
            if sucesso:
                self.adicionar_log("\n" + "="*60, "SUCCESS")
                self.adicionar_log(f"✅ Relatório gerado com sucesso!", "SUCCESS")
                self.adicionar_log(f"📄 Arquivo: {os.path.basename(arquivo_destino)}", "SUCCESS")
                self.adicionar_log(f"📁 Local: {os.path.dirname(arquivo_destino)}", "SUCCESS")
                self.adicionar_log("="*60, "SUCCESS")
                
                # Perguntar se quer abrir
                resposta = messagebox.askyesno(
                    "Sucesso",
                    f"Relatório gerado com sucesso!\n\n"
                    f"Arquivo: {os.path.basename(arquivo_destino)}\n"
                    f"XMLs processados: {len(lista_xmls)}\n\n"
                    f"Deseja abrir o arquivo agora?"
                )
                
                if resposta:
                    try:
                        if os.name == 'nt':  # Windows
                            os.startfile(arquivo_destino)
                        elif os.name == 'posix':  # Linux/Mac
                            import subprocess
                            subprocess.call(['xdg-open', arquivo_destino])
                    except Exception as e:
                        self.adicionar_log(f"⚠️ Não foi possível abrir o arquivo: {e}", "WARNING")
            else:
                self.adicionar_log("\n❌ Erro ao gerar relatório", "ERROR")
                messagebox.showerror("Erro", "Erro ao gerar relatório.\nVerifique o log para detalhes.")
            
        except Exception as e:
            self.adicionar_log(f"\n❌ ERRO: {e}", "ERROR")
            logger.error(f"Erro _gerar_relatorio_xmls_selecionados: {e}\n{traceback.format_exc()}")
            messagebox.showerror("Erro", f"Erro ao gerar relatório:\n{e}")
    
    def reset_ui(self):
        """Reseta interface após execução"""
        def _reset():
            self.btn_run.config(state="normal")
            self.btn_cancelar_exec.config(state="disabled")
            self.cancelar_flag = False
        
        self.root.after(0, _reset)


# ============================================================================
# EXECUÇÃO PRINCIPAL
# ============================================================================

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = NFSeDownloaderApp(root)
        root.mainloop()
    except Exception as e:
        logger.critical(f"Erro fatal na aplicação:\n{traceback.format_exc()}")
        messagebox.showerror(
            "Erro Fatal",
            f"Erro ao iniciar aplicação:\n{e}\n\nVerifique o arquivo de log."
        )